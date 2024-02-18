import openpyxl
from openpyxl.styles import Protection

def lock_cells(input_file, output_file):
    # Load the input workbook
    workbook = openpyxl.load_workbook(input_file)

    # Iterate through all sheets in the workbook
    for sheet in workbook.sheetnames:
        # Get the active sheet
        worksheet = workbook[sheet]

        # Iterate through all cells in the sheet
        for row in worksheet.iter_rows():
            for cell in row:
                # Lock the cell
                cell.protection = Protection(locked=True)

    # Save the modified workbook as the output
    workbook.save(output_file)

# Usage example
input_file = 'input.xlsx'
output_file = 'output.xlsx'
lock_cells(input_file, output_file)