from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Protection
import numpy as np


def printout(aw, Number_of_COs, start_row):
    # ===========================================================================================================     
            
    start_column=4

    column=start_column
    row=start_row
    # Merging cells dynamically based on row and column number
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)

    # Setting value, font, and alignment for the merged cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Code"
    aw[cell_reference].font = Font(bold=True)
    aw[cell_reference].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)	#wrapping text

    # Applying border to the cells in the range
    for r in aw.iter_rows(min_row=row, max_row=row+2, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    aw.merge_cells(start_row=row+3, end_row=row+3+Number_of_COs-1, start_column=column,  end_column=column)
    Subject_Code_cell_reference = f"{get_column_letter(column)}{row+3}"
    aw[Subject_Code_cell_reference].font = Font(bold=True)
    aw[Subject_Code_cell_reference].alignment = Alignment(horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    aw[Subject_Code_cell_reference].fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
    for r in aw.iter_rows(min_row=row+3, max_row=row+3+Number_of_COs-1, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    # Merging cells for "Course Name"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)

    # Setting value, font, and alignment for "Course Name" cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Name"
    aw[cell_reference].font = Font(bold=True)
    aw[cell_reference].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    

    # Applying border to the cells in the "Course Name" range
    for r in aw.iter_rows(min_row=row, max_row=row+2, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))    
    aw.merge_cells(start_row=row+3, end_row=row+3+Number_of_COs-1, start_column=column,  end_column=column)
    Subject_Name_cell_reference = f"{get_column_letter(column)}{row+3}"
    aw[Subject_Name_cell_reference].font = Font(bold=True)
    aw[Subject_Name_cell_reference].alignment = Alignment(horizontal='center', vertical='center', textRotation=90, wrap_text=True)
    aw[Subject_Name_cell_reference].fill = PatternFill(start_color='1ed760', end_color='1ed760', fill_type='solid')
    for r in aw.iter_rows(min_row=row+3, max_row=row+3+Number_of_COs-1, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1

    #============================================================================================================            
    interval=17

    # Merging cells for "COs"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)

    # Setting value, font, and alignment for "COs" cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "COs"
    aw[cell_reference].font = Font(bold=True)
    aw[cell_reference].alignment = Alignment(horizontal='center', vertical='center')

    # Applying border to the cells in the "COs" range
    for r in aw.iter_rows(min_row=row, max_row=row+2, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            

    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"] = f"CO{numco+1}"
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        if numco%2==0:
            aw[f"{get_column_letter(column)}{row+3+numco}"].fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1          
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "End Semester Examination" cell
    ese_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ese_cell_reference] = "End Semester Examination"
    aw[ese_cell_reference].font = Font(bold=True)
    aw[ese_cell_reference].alignment = Alignment(horizontal='center', vertical='center')

    # Applying border to the cells in the "End Semester Examination" range
    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            
    
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "(SEE)*" cell
    see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[see_cell_reference] = "(SEE)*"
    aw[see_cell_reference].font = Font(bold=True)
    aw[see_cell_reference].alignment = Alignment(horizontal='center', vertical='center')

    # Applying border to the cells in the "(SEE)*" range
    for r in aw.iter_rows(min_row=row+1, max_row=row+1, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))


    #============================================================================================================
    
    start_row_ca_data=Number_of_COs+8+Number_of_COs+3+4
    start_col_ca_data=4+3
            
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw[attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw[level_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[level_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Internal Examination" cell
    ie_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ie_cell_reference] = "Internal Examination"
    aw[ie_cell_reference].font = Font(bold=True)
    aw[ie_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================

    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "(CIE)*" cell
    cie_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_cell_reference] = "(CIE)*"
    aw[cie_cell_reference].font = Font(bold=True)
    aw[cie_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row+1, max_row=row+1, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================        
    
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw[attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
   
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw[level_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[level_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Direct" cell
    direct_cell_reference = f"{get_column_letter(column)}{row}"
    aw[direct_cell_reference] = "Direct"
    aw[direct_cell_reference].font = Font(bold=True)
    aw[direct_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "CIE + SEE" cell
    cie_see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_see_cell_reference] = f"=B15 & \" % of CIE + \" & B16 & \" % of SEE\""
    aw[cie_see_cell_reference].font = Font(bold=True)
    aw[cie_see_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row+1, max_row=row+1, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw[attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
        
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw[level_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[level_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(column)}{row+3+numco}"].fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Indirect" cell
    indirect_cell_reference = f"{get_column_letter(column)}{row}"
    aw[indirect_cell_reference] = "Indirect"
    aw[indirect_cell_reference].font = Font(bold=True)
    aw[indirect_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row, max_row=row+1, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw[attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12    
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw[level_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[level_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(column)}{row+3+numco}"].fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Total Course Attainment" cell
    total_course_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[total_course_attainment_cell_reference] = "Total Course Attainment"
    aw[total_course_attainment_cell_reference].font = Font(bold=True)
    aw[total_course_attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Direct + Indirect" cell
    direct_indirect_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[direct_indirect_cell_reference] = f"=B17 & \" % of Direct + \" & B18 & \" % of Indirect\""
    aw[direct_indirect_cell_reference].font = Font(bold=True)
    aw[direct_indirect_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row+1, max_row=row+1, min_col=column, max_col=column+1):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw[attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font= Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20                            
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw[level_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[level_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(column)}{row+3+numco}"].fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    target_cell_reference = f"{get_column_letter(column)}{row}"
    aw[target_cell_reference] = "Target"
    aw[target_cell_reference].font = Font(bold=True)
    aw[target_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column):
        for c in r:
            c.border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    #============================================================================================================
    percentage_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[percentage_cell_reference] = "(%)"
    aw[percentage_cell_reference].font = Font(bold=True)
    aw[percentage_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[percentage_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"] = f"=B19"
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(column)}{row+3+numco}"].fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    final_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[final_attainment_cell_reference] = "Final Attainment"
    aw[final_attainment_cell_reference].font = Font(bold=True)
    aw[final_attainment_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[final_attainment_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    #============================================================================================================
    yesno_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[yesno_cell_reference] = "Yes/No"
    aw[yesno_cell_reference].font = Font(bold=True)
    aw[yesno_cell_reference].alignment = Alignment(horizontal='center', vertical='center')
    aw[yesno_cell_reference].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    for numco in range(Number_of_COs):
        aw[f"{get_column_letter(column)}{row+3+numco}"] = f'=IF({get_column_letter(column-3)}{row+3+numco}>=B19,"Yes","No")'
        aw[f"{get_column_letter(column)}{row+3+numco}"].font = Font(bold=True)
        aw[f"{get_column_letter(column)}{row+3+numco}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(column)}{row+3+numco}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
        
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20
    column+=1
    #============================================================================================================
    

    column=start_column+3
    row=start_row+2

    for r in aw.iter_rows(min_row=row, max_row=row, min_col=column, max_col=aw.max_column):
        for c in r:
            c.fill = PatternFill(start_color='8db4e2', end_color='8db4e2', fill_type='solid')

    
    return aw