from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Protection 
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import FormulaRule


def CO_PO_Table(data,aw):
    #merge cells depending on number of POs
    aw.merge_cells(start_row=1, start_column=4, end_row=1, end_column=12+5+1+3)
    aw['D1']="CO-PO Mapping"
    aw['D1'].font = Font(bold=True)
    aw['D1'].alignment = Alignment(horizontal='center', vertical='center')

    aw["D2"]="COs\\POs"
    aw["D2"].font = Font(bold=True)


    for co in range(1,data["Number_of_COs"]+1):
        aw[f"D{co+2}"]=f"CO{co}"
        aw[f"D{co+2}"].font = Font(bold=True)
    for po in range(1,12+1):
        aw[f"{get_column_letter(po+4)}2"]=f"PO{po}   "
        aw[f"{get_column_letter(po+4)}2"].font = Font(bold=True)
        aw[f"{get_column_letter(po+4)}2"].alignment = Alignment(horizontal='center', vertical='center')
    for pso in range(1,6):
        aw[f"{get_column_letter(12+4+pso)}2"]=f"PSO{pso}"
        aw[f"{get_column_letter(12+4+pso)}2"].font = Font(bold=True)
        aw[f"{get_column_letter(12+4+pso)}2"].alignment = Alignment(horizontal='center', vertical='center')
    
    
    #for columns 4 to 4+12+5 set width to 13
    for col in range(4,4+12+5):
        aw.column_dimensions[f"{get_column_letter(col)}"].width = 13

    #make it into a table
    tab = Table(displayName=f"CO_PO", ref=f"D2:{get_column_letter(12+4+5)}{data['Number_of_COs']+2}")  #create table
    style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)  #set style
    tab.tableStyleInfo = style                                                                                                                   #set style
    aw.add_table(tab)

   

            
    #set conditional formatting for empty cells
        
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    for co in range(1,data["Number_of_COs"]+1):
        for po in range(1,12+1):
            cell_coordinate = f"{get_column_letter(po+4)}{co+2}"
            aw.conditional_formatting.add(
                cell_coordinate,
                FormulaRule(formula=[f'ISBLANK({cell_coordinate})'], stopIfTrue=False, fill=pink_fill)
            )
            
        for pso in range(1,6):
            cell_coordinate = f"{get_column_letter(12+4+pso)}{co+2}"
            aw.conditional_formatting.add(
                cell_coordinate,
                FormulaRule(formula=[f'ISBLANK({cell_coordinate})'], stopIfTrue=False, fill=pink_fill)
            )
         
     # Create a data validation object for numbers between 0 and 100
    number_validation = DataValidation(type="whole", operator="between", formula1=0, formula2=3)
    number_validation.error = 'You must enter a number 1, 2 or 3'
    number_validation.errorTitle = 'Invalid Entry'
    number_validation.showErrorMessage = True

    # Apply this validation to the specified cells
    for co in range(1,data["Number_of_COs"]+1):
        for po in range(1,12+1):
            cell_coordinate = f"{get_column_letter(po+4)}{co+2}"
            aw.add_data_validation(number_validation)
            number_validation.add(aw[cell_coordinate])
        for pso in range(1,6):
            cell_coordinate = f"{get_column_letter(12+4+pso)}{co+2}"
            aw.add_data_validation(number_validation)
            number_validation.add(aw[cell_coordinate])




    green_fill=PatternFill(start_color="5e9955", end_color="5e9955", fill_type='solid')                                                                                                                     #add table to workshee
    # Lock all cells by default
    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)

    aw.protection.sheet = True
    #unlocked cells B9 to B19
    for i in range(9,20):
        if i!=16 and i!=18:
            aw[f'B{i}'].protection = Protection(locked=False)
            #aw[f'B{i}'].fill = green_fill

    #unlocked cells E3 to U[2+Number_of_COs]
    for row in range(3,3+data['Number_of_COs']):
        for col in range(5,22):
            aw.cell(row=row,column=col).protection = Protection(locked=False)
            #aw.cell(row=row,column=col).fill = green_fill

    #unlocked cells E[2+Number_of_COs+4] to E[2+Number_of_COs+4+Number_of_COs]
    for row in range(2+data['Number_of_COs']+5, 2+data['Number_of_COs']+5+data['Number_of_COs']):
        aw.cell(row=row,column=5).protection = Protection(locked=False)
        #aw.cell(row=row,column=5).fill = green_fill



    return aw