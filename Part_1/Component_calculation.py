from openpyxl import *
from .utils import copy_range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Protection



def Component_calculation(data,Component_Details,aw,component_type):
    #get the value from other sheets and put it in this sheet as a formula dynamically by referencing the cell
    start_column=1
    component_num=2
    components_len=0
    for component_name in Component_Details.keys():
        if component_name[-1]==component_type:
            aw[f'{get_column_letter(start_column)}1']=component_name
            aw.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column+data['Number_of_COs']-1)
            aw[f'{get_column_letter(start_column)}1'].alignment = Alignment(horizontal='center')
            aw[f'{get_column_letter(start_column)}1'].font = Font(bold=True)

            #reference a cell from another sheet
            for nco in range(1,data['Number_of_COs']+1):
                aw[f'{get_column_letter(start_column+nco-1)}2']=f"CO{nco}"
                #bold and center align the cell
                aw[f'{get_column_letter(start_column+nco-1)}2'].alignment = Alignment(horizontal='center')
                aw[f'{get_column_letter(start_column+nco-1)}2'].font = Font(bold=True)

                aw[f'{get_column_letter(start_column+nco-1)}3']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}3"
                #bold and center align the cell
                aw[f'{get_column_letter(start_column+nco-1)}3'].alignment = Alignment(horizontal='center')

                aw[f'{get_column_letter(start_column+nco-1)}4']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}4"
                #bold and center align the cell
                aw[f'{get_column_letter(start_column+nco-1)}4'].alignment = Alignment(horizontal='center')
            
            for nco in range(1,data['Number_of_COs']+1):
                aw[f'{get_column_letter(start_column+nco-1)}6']=f"CO{nco}"
                #bold and center align the cell
                aw[f'{get_column_letter(start_column+nco-1)}6'].alignment = Alignment(horizontal='center')
                aw[f'{get_column_letter(start_column+nco-1)}6'].font = Font(bold=True)

                for nstudents in range(1,data['Number_of_Students']+1):
                    aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}']=f"='{component_name}'!{get_column_letter(2+Component_Details[component_name]+1+nco)}{10+nstudents}"
                    aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}'].alignment = Alignment(horizontal='center')
                    aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}'].font = Font(bold=True)

            table_range = f"{get_column_letter(start_column)}2:{get_column_letter(start_column+data['Number_of_COs']-1)}4"
            tab = Table(displayName=f"{component_name}_ComponentData", ref=table_range)
            style = TableStyleInfo(name=f"TableStyleMedium{component_num}", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            aw.add_table(tab)

            table_range = f"{get_column_letter(start_column)}6:{get_column_letter(start_column+data['Number_of_COs']-1)}{6+data['Number_of_Students']}"
            tab = Table(displayName=f"{component_name}_StudentMarks", ref=table_range)
            style = TableStyleInfo(name=f"TableStyleMedium{component_num}", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            aw.add_table(tab)

            start_column+=data['Number_of_COs']+1 
            component_num+=1
            components_len+=1
    #make a column of rows black to demarcate the components
    #set width of columns
    aw.column_dimensions[f'{get_column_letter(start_column)}'].width = 2.5
    for i in range(1,data['Number_of_Students']+12):
        aw[f'{get_column_letter(start_column)}{i}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    start_column+=2
    #set header to Combined Components
    aw[f'{get_column_letter(start_column)}1']="Combined Components table"
    aw.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=start_column+data['Number_of_COs']-1)
    aw[f'{get_column_letter(start_column)}1'].alignment = Alignment(horizontal='center')
    aw[f'{get_column_letter(start_column)}1'].font = Font(bold=True)

    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column+nco-1)}2']=f"CO{nco}"
        #bold and center align the cell
        aw[f'{get_column_letter(start_column+nco-1)}2'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column+nco-1)}2'].font = Font(bold=True)
        aw[f'{get_column_letter(start_column+nco-1)}2'].font = Font(color='FFFFFF')    

        corr_column=nco
        formula="=SUM("
        for _ in range(1,components_len+1):
            formula+=f"{get_column_letter(corr_column)}3,"
            corr_column+=data['Number_of_COs']+1
        formula=formula[:-1]
        formula+=")"
        aw[f'{get_column_letter(start_column+nco-1)}3']=formula
        aw[f'{get_column_letter(start_column+nco-1)}3'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column+nco-1)}3'].font = Font(bold=True)

        corr_column=nco
        formula="=SUM("
        for _ in range(1,components_len+1):
            formula+=f"{get_column_letter(corr_column)}4,"
            corr_column+=data['Number_of_COs']+1
        formula=formula[:-1]
        formula+=")"
        aw[f'{get_column_letter(start_column+nco-1)}4']=formula
        aw[f'{get_column_letter(start_column+nco-1)}4'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column+nco-1)}4'].font = Font(bold=True)

    table_range = f"{get_column_letter(start_column)}2:{get_column_letter(start_column+data['Number_of_COs']-1)}4"
    tab = Table(displayName=f"Combined_ComponentData_{component_type}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{1}", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    aw.add_table(tab)

    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column+nco-1)}6']=f"CO{nco}"
        #bold and center align the cell
        aw[f'{get_column_letter(start_column+nco-1)}6'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column+nco-1)}6'].font = Font(bold=True)
        aw[f'{get_column_letter(start_column+nco-1)}6'].font = Font(color='FFFFFF')

        for nstudents in range(1,data['Number_of_Students']+1):
            corr_column=nco
            formula="=SUM("
            for _ in range(1,components_len+1):
                formula+=f"{get_column_letter(corr_column)}{6+nstudents},"
                corr_column+=data['Number_of_COs']+1
            formula=formula[:-1]
            formula+=")"
            aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}']=formula
            aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}'].alignment = Alignment(horizontal='center')
            aw[f'{get_column_letter(start_column+nco-1)}{6+nstudents}'].font = Font(bold=True)

    table_range = f"{get_column_letter(start_column)}6:{get_column_letter(start_column+data['Number_of_COs']-1)}{6+data['Number_of_Students']}"
    tab = Table(displayName=f"Combined_StudentMarks_{component_type}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{1}", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    aw.add_table(tab)

    aw.column_dimensions[f'{get_column_letter(start_column-1)}'].width = 14.3
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}']="CO"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}'].alignment = Alignment(horizontal='center')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}'].font = Font(bold=True)
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+8}'].font = Font(color='FFFFFF')

    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}']="CO%"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}'].alignment = Alignment(horizontal='center')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}'].font = Font(bold=True)
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+9}'].font = Font(color='FFFFFF')

    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}']="Total students"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}'].alignment = Alignment(horizontal='center')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}'].font = Font(bold=True)
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+10}'].font = Font(color='FFFFFF')

    if component_type=="I":
            aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}']="I_attainment %"
    else:
        aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}']="E_attainment %"
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}'].alignment = Alignment(horizontal='center')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}'].font = Font(bold=True)
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    aw[f'{get_column_letter(start_column-1)}{data["Number_of_Students"]+11}'].font = Font(color='FFFFFF')

    for nco in range(1,data['Number_of_COs']+1):
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}']=f"CO{nco}"
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}'].font = Font(bold=True)
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+8}'].font = Font(color='FFFFFF')

        # Let's say the criterion is to count cells greater than or equal to the value in cell at (column, row) = (start_column-1+nco, 4)
        range_start = f"{get_column_letter(start_column-1+nco)}7"
        range_end = f"{get_column_letter(start_column-1+nco)}{6+data['Number_of_Students']}"
        range_string = f"{range_start}:{range_end}"

        # Building the COUNTIF formula with an embedded IF condition
        criteria_cell = f"{get_column_letter(start_column-1+nco)}4"
        criteria = f'">=" & {criteria_cell}'

        # The IF condition checks if the sum of the range is greater than zero (indicating non-zero values are present)
        formula = f'=IF(SUM({range_string}) > 0, COUNTIF({range_string}, {criteria}), "")'  # Empty string or "N/A" as placeholder

        # Write the formula to the cell
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+9}'] = formula
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+9}'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+9}'].font = Font(bold=True)

        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+10}']=data['Number_of_Students']
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+10}'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+10}'].font = Font(bold=True)

        #aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+11}']=f"={get_column_letter(start_column-1+nco)}{data['Number_of_Students']+9}/{get_column_letter(start_column-1+nco)}{data['Number_of_Students']+10}*100"
        num_students = data['Number_of_Students']
        cell_position_1 = f"{get_column_letter(start_column-1+nco)}{num_students+9}"
        cell_position_2 = f"{get_column_letter(start_column-1+nco)}{num_students+10}"
        target_cell_position = f"{get_column_letter(start_column-1+nco)}{num_students+11}"

        # Construct the formula
        formula = f'=IF(SUM({range_string}) > 0, {cell_position_1}/{cell_position_2}*100, "0")'

        # Assign the formula to the cell
        aw[target_cell_position] = formula
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+11}'].alignment = Alignment(horizontal='center')
        aw[f'{get_column_letter(start_column-1+nco)}{data["Number_of_Students"]+11}'].font = Font(bold=True)

    #make it a table
    table_range = f"{get_column_letter(start_column)}{data['Number_of_Students']+8}:{get_column_letter(start_column+data['Number_of_COs']-1)}{data['Number_of_Students']+11}"
    tab = Table(displayName=f"Final_attainment_{component_type}", ref=table_range)
    style = TableStyleInfo(name=f"TableStyleMedium{1}", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    aw.add_table(tab)

    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    aw.protection.sheet = True

    return aw

