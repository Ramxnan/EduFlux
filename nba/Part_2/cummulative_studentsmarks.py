from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.styles import Font                                       #import font and alignment from openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment


def cummulative_studentmarks(data, component_name, component_details, aw):
    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(component_details+3+cno)}10'] = f'CO{cno}'
        aw[f'{get_column_letter(component_details+3+cno)}10'].font = Font(bold=True, color='FFFFFF')
        aw[f'{get_column_letter(component_details+3+cno)}10'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        
        co_name = f'{data["Subject_Code"]}_CO{cno}'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(component_details + 2)}6'

        for nstudents in range(1, data["Number_of_Students"]+1):
            sum_range_start_marks = f'C{10+nstudents}'
            sum_range_end_marks = f'{get_column_letter(component_details + 2)}{10+nstudents}'
            aw[f'{get_column_letter(component_details+3+cno)}{10+nstudents}'] = f'=SUMIFS({sum_range_start_marks}:{sum_range_end_marks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'


    for cno in range(1, data["Number_of_COs"]+1):
        for nstudents in range(data["Number_of_Students"]+1):
            aw[f'{get_column_letter(component_details+3+cno)}{10+nstudents}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            aw[f'{get_column_letter(component_details+3+cno)}{10+nstudents}'].alignment = Alignment(horizontal='center', vertical='center')
    

    return aw


            
    