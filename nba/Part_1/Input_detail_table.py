from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Protection 
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from .utils import colour_table_Input_Details, cellstyle, cellstyle_range

#Input details
def input_detail(data,Component_Details,aw):  #function to input details
    aw.merge_cells('A1:B1')
    aw['A1']="Constants"
    cellstyle_range(aw['A1:B1'], bold=True, alignment=True, border=True, fill="ffe74e")

    startrow=2
    for key, value in data.items():
        aw[f'A{startrow}']=key
        aw[f'B{startrow}']=value
        startrow+=1

    cellstyle_range(aw[f'A2:B{startrow-1}'], border=True, alignment=True, bold=True, alternate=['b7dee8', 'daeef3'])
    startrow+=1

    aw.merge_cells(f'A{startrow}:B{startrow}')
    aw[f'A{startrow}']="Variables"
    cellstyle_range(aw[f'A{startrow}:B{startrow}'], bold=True, alignment=True, border=True, fill="ffe74e")
    
    aw['A14']="Default Threshold %"
    aw['A15']="Internal %"
    aw['A16']="External %"
    aw['B16']=f'=100-B15'
    aw['A17']="Direct %"
    aw['A18']="Indirect %"
    aw['B18']=f'=100-B17'
    aw['A19']="Target CO Attainment %"

    cellstyle_range(aw[f'A14:B19'], border=True, alignment=True, bold=True, alternate=['b7dee8', 'daeef3'])

    # =================================================================================================================================================================
    aw['A22']="Component Details"
    aw['B22']="Number of Questions"
    cellstyle_range(aw['A22:B22'], bold=True, alignment=True, border=True, font_color="FFFFFF")

    startrow=23
    for key, value in Component_Details.items():
        aw[f'A{startrow}']=key
        aw[f'B{startrow}']=value
        cellstyle_range(aw[f'A{startrow}:B{startrow}'], alignment=True, border=True, bold=True)
        startrow+=1
      
    #make a table
    tab = Table(displayName=f"{data['Section']}_Component_Details", ref=f"A22:B{aw.max_row}")
    style = TableStyleInfo(name="TableStyleMedium15", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    aw.add_table(tab)
    
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
    #set conditional formatting for B9 to B19 such that if its empty, it will be highlighted pink
    for startrow in range(14,20):
        if startrow!=16 and startrow!=18:
            aw.conditional_formatting.add(
                f'B{startrow}',
                FormulaRule(formula=[f'ISBLANK(B{startrow})'], stopIfTrue=False, fill=pink_fill))
            aw.conditional_formatting.add(
                f'B{startrow}',
                #greater than 100 or less than 0
                FormulaRule(formula=[f'OR(B{startrow}>100,B{startrow}<0)'], stopIfTrue=False, fill=red_fill))

    colour_table_Input_Details(aw)
    
    return aw  
            