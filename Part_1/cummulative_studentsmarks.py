from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Font                                       #import font and alignment from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import CellIsRule
from .utils import colour_table


def cummulative_studentmarks(data, component_name, component_details, aw):
    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(component_details+3+cno)}10'] = f'CO{cno}'
        
        co_name = f'{data["Subject_Code"]}_CO{cno}'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(component_details + 2)}6'

        for nstudents in range(1, data["Number_of_Students"]+1):
            sum_range_start_marks = f'C{10+nstudents}'
            sum_range_end_marks = f'{get_column_letter(component_details + 2)}{10+nstudents}'
            aw[f'{get_column_letter(component_details+3+cno)}{10+nstudents}'] = f'=SUMIFS({sum_range_start_marks}:{sum_range_end_marks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'

    #make it a table
    table_range = f"{get_column_letter(component_details + 4)}10:{get_column_letter(component_details + 3 + data['Number_of_COs'])}{10+data['Number_of_Students']}"
    tab = Table(displayName=f"cummulative_studentmarks_{component_name}", ref=table_range)
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style
    aw.add_table(tab)

    colour_table(aw, data)

    return aw


            
    