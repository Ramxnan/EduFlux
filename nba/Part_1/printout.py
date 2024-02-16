from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.utils import  get_column_letter
from .utils import adjust_width, cellstyle_range, cellstyle
from openpyxl.styles import Protection
from .Input_Details import input_detail

def printout(aw, data,start_row,copy=False):

    #============================================================================================================
    if not copy:
        aw = input_detail(data, {}, aw, copy=True)
        adjust_width(aw)


    # ===========================================================================================================     
            
    start_column=4

    column=start_column
    row=start_row
    # Merging cells dynamically based on row and column number
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)
    # Setting value, font, and alignment for the merged cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Code"
    
    aw.merge_cells(start_row=row+3, end_row=row+3+data['Number_of_COs']-1, start_column=column,  end_column=column)
    Subject_Code_cell_reference = f"{get_column_letter(column)}{row+3}"
    aw[Subject_Code_cell_reference] = data["Subject_Code"]
    cellstyle(aw[Subject_Code_cell_reference], text_rotation=90, fill="ffff00")
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    # Merging cells for "Course Name"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)

    # Setting value, font, and alignment for "Course Name" cell
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "Course Name"
     
    aw.merge_cells(start_row=row+3, end_row=row+3+data['Number_of_COs']-1, start_column=column,  end_column=column)
    Subject_Name_cell_reference = f"{get_column_letter(column)}{row+3}"
    aw[Subject_Name_cell_reference] = data["Subject_Name"]
    cellstyle(aw[Subject_Name_cell_reference], text_rotation=90, fill="1ed760")

    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1

    #============================================================================================================            
    interval=17

    # Merging cells for "COs"
    aw.merge_cells(start_row=row, end_row=row+2, start_column=column,  end_column=column)
    cell_reference = f"{get_column_letter(column)}{row}"
    aw[cell_reference] = "COs"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1          
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    ese_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ese_cell_reference] = "End Semester Examination"
    
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[see_cell_reference] = "(SEE)*"


    #============================================================================================================
    
    start_row_ca_data=data["Number_of_COs"]+8+data["Number_of_COs"]+3+4
    start_col_ca_data=4+3
            
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    
    for numco in range(data['Number_of_COs']):
        aw[f"{get_column_letter(column)}{row+3+numco}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+1)}{start_row_ca_data+(numco*interval)}"
    
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    ie_cell_reference = f"{get_column_letter(column)}{row}"
    aw[ie_cell_reference] = "Internal Examination"
    
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    cie_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_cell_reference] = "(CIE)*"
    
    #============================================================================================================        
    
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
   
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"  
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1

    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "Direct" cell
    direct_cell_reference = f"{get_column_letter(column)}{row}"
    aw[direct_cell_reference] = "Direct"
   
    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)

    # Setting value, font, and alignment for "CIE + SEE" cell
    cie_see_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[cie_see_cell_reference] = f"=B15 & \" % of CIE + \" & B16 & \" % of SEE\""

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"  
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
        
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row+1, start_column=column, end_column=column+1)
    indirect_cell_reference = f"{get_column_letter(column)}{row}"
    aw[indirect_cell_reference] = "Indirect"

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 12    
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    aw.merge_cells(start_row=row, end_row=row, start_column=column, end_column=column+1)
    total_course_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[total_course_attainment_cell_reference] = "Total Course Attainment"

    #============================================================================================================
    aw.merge_cells(start_row=row+1, end_row=row+1, start_column=column, end_column=column+1)
    direct_indirect_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[direct_indirect_cell_reference] = f"=B17 & \" % of Direct + \" & B18 & \" % of Indirect\""

    #============================================================================================================
    attainment_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[attainment_cell_reference] = "Attainment"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20                            
    column+=1
    #============================================================================================================
    level_cell_reference = f"{get_column_letter(column)}{row+2}"
    aw[level_cell_reference] = "Level"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    target_cell_reference = f"{get_column_letter(column)}{row}"
    aw[target_cell_reference] = "Target"
    
    #============================================================================================================
    percentage_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[percentage_cell_reference] = "(%)"
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 8.43
    column+=1
    #============================================================================================================
    final_attainment_cell_reference = f"{get_column_letter(column)}{row}"
    aw[final_attainment_cell_reference] = "Final Attainment"

    #============================================================================================================
    yesno_cell_reference = f"{get_column_letter(column)}{row+1}"
    aw[yesno_cell_reference] = "Yes/No"
    aw[yesno_cell_reference].font = Font(bold=True)
    aw.column_dimensions[f"{get_column_letter(column)}"].width = 20
    column+=1
    #============================================================================================================

    #Printing all the data in the excel sheet
    start_row_ca_data=data["Number_of_COs"]+8+data["Number_of_COs"]+3+4
    start_col_ca_data=4+3
    start_column=4
    start_row=4
    aw[f"{get_column_letter(start_column)}{start_row}"] = data['Subject_Code']
    aw[f"{get_column_letter(start_column+1)}{start_row}"] = data['Subject_Name']
    for numco in range(data['Number_of_COs']):
        aw[f"{get_column_letter(start_column+2)}{start_row}"] = f"CO{numco+1}"
        aw[f"{get_column_letter(start_column+3)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data)}{start_row_ca_data+(numco*interval)}"
        aw[f"{get_column_letter(start_column+4)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+1)}{start_row_ca_data+(numco*interval)}"
        cellstyle(aw[f"{get_column_letter(start_column+4)}{start_row}"], fill="fde9d9")
        aw[f"{get_column_letter(start_column+5)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+2)}{start_row_ca_data+(numco*interval)}"
        aw[f"{get_column_letter(start_column+6)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+3)}{start_row_ca_data+(numco*interval)}"
        cellstyle(aw[f"{get_column_letter(start_column+6)}{start_row}"], fill="fde9d9")
        aw[f"{get_column_letter(start_column+7)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+4)}{start_row_ca_data+(numco*interval)}"
        aw[f"{get_column_letter(start_column+8)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+5)}{start_row_ca_data+(numco*interval)}"
        cellstyle(aw[f"{get_column_letter(start_column+8)}{start_row}"], fill="fde9d9")
        aw[f"{get_column_letter(start_column+9)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+6)}{start_row_ca_data+(numco*interval)}"
        aw[f"{get_column_letter(start_column+10)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+7)}{start_row_ca_data+(numco*interval)}"
        cellstyle(aw[f"{get_column_letter(start_column+10)}{start_row}"], fill="fde9d9")
        aw[f"{get_column_letter(start_column+11)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+8)}{start_row_ca_data+(numco*interval)}"
        aw[f"{get_column_letter(start_column+12)}{start_row}"] = f"={data['Section']}_Course_Attainment!{get_column_letter(start_col_ca_data+9)}{start_row_ca_data+(numco*interval)}"
        cellstyle(aw[f"{get_column_letter(start_column+12)}{start_row}"], fill="fde9d9")
        aw[f"{get_column_letter(start_column+13)}{start_row}"] = f"=B19"
        cellstyle(aw[f"{get_column_letter(start_column+13)}{start_row}"], fill="ffff00")
        aw[f"{get_column_letter(start_column+14)}{start_row}"] = f'=IF({get_column_letter(start_column+11)}{start_row}>=B19,"Yes","No")'
        start_row+=1


   
    #============================================================================================================

    cellstyle_range(aw[f"D1:R{3+data['Number_of_COs']}"],border=True, bold=True, alignment=True, wrap_text=True)
    cellstyle_range(aw[f"G{3}:R{3}"],border=True, bold=True, alignment=True, wrap_text=True, fill="8db4e2")
    return aw