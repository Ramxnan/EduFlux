from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl import Workbook                                                         #import workbook from openpyxl
from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Protection 
from openpyxl.worksheet.datavalidation import DataValidation
from .utils import adjust_width

def studentmarks(data,key, Component_details,aw):
    aw.merge_cells(f'B9:{get_column_letter(Component_details+2)}9')
    aw["B9"]="Marks obtained"
    aw["B9"].font = Font(bold=True)

    aw["A10"]="Roll No."
    aw["A10"].font = Font(bold=True)
    #set column A width to 13
    aw.column_dimensions['A'].width = 13


    aw["B10"]="Name"
    aw["B10"].font = Font(bold=True)

    for qno in range(1,Component_details+1):
        aw[get_column_letter(qno+2)+'10']=f"Q{qno}"
        aw[get_column_letter(qno+2)+'10'].font = Font(bold=True)

    #conditional formatting
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    light_red_fill = PatternFill(start_color="FF5E5E", end_color="FF5E5E", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD9A46F", end_color="FFD9A46F", fill_type="solid")

    # Apply conditional formatting to each question's column header
    for qno in range(1, Component_details + 1):
        column_letter = get_column_letter(qno + 2)
        header_cell = f"{column_letter}10"
        data_range = f"{column_letter}11:{column_letter}{10 + data['Number_of_Students']}"
        threshold_cell = f"{column_letter}4"
        max_marks_cell = f"${column_letter}$3"  # Assuming max marks is in row 3
        all_range = f"A11:{get_column_letter(Component_details + 2)}{10 + data['Number_of_Students']}"
        # Apply conditional formatting to the header based on the cells below it
        # The formula checks if all cells in the range below the header are below the threshold
        formula = f"COUNTIF({data_range}, \">=\"&{threshold_cell})=0"
        
        # Since the COUNTIF will return count of all cells not less than the threshold, we only apply
        # the format if the result of COUNTIF is 0, meaning all filled cells are below the threshold
        aw.conditional_formatting.add(
            header_cell,
            FormulaRule(formula=[formula], stopIfTrue=False, fill=yellow_fill)
        )
       
        #Highlight empty cells
        aw.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'ISBLANK({column_letter}11)'], stopIfTrue=False, fill=pink_fill)
        )

        #Highlight cells with value greater than max marks
        aw.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'{column_letter}11>{max_marks_cell}'], stopIfTrue=False, fill=light_red_fill)
        )

        #highlight roll no. and name cells with pink if empty
        aw.conditional_formatting.add(
            f"A11:A{10 + data['Number_of_Students']}",
            FormulaRule(formula=[f'ISBLANK(A11)'], stopIfTrue=False, fill=pink_fill)
        )
        aw.conditional_formatting.add(
            f"B11:B{10 + data['Number_of_Students']}",
            FormulaRule(formula=[f'ISBLANK(B11)'], stopIfTrue=False, fill=pink_fill)
        )

    #unprotect all range
    for row in aw.iter_rows(min_row=11, max_row=10+data['Number_of_Students'], min_col=1, max_col=Component_details+2):
        for cell in row:
            cell.protection = Protection(locked=False)

    #set cell A10 and B10 to colour #4bacc6
    aw['A10'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    #set border around the cell A10 and B10 to thin and black
    aw['A10'].border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
    aw['B10'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    #set border around the cell A10 and B10 to thin and black
    aw['B10'].border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
    
    #set cells A11 till B[10+Number_of_Students] boreder to purple and thin
    for row in aw.iter_rows(min_row=11, max_row=10+data['Number_of_Students'], min_col=1, max_col=2+Component_details):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            
    adjust_width(aw)
    aw.column_dimensions['A'].width = 20
    aw.column_dimensions['B'].width = 30


    table_range = f"C10:{get_column_letter(Component_details + 2)}{data['Number_of_Students'] + 10}"
    tab = Table(displayName=f"studentmarks_{key}", ref=table_range)
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style
    aw.add_table(tab)

    

    


    return aw