from openpyxl import Workbook
from .utils import adjust_width
from .Input_Details import input_detail, indirect_co_assessment, CO_PO_Table
from .Component_values import qn_co_mm_btl, studentmarks
from .Cummulative_Component_Values import cummulative_qn_co_mm_btl, cummulative_studentmarks
from .Component_calculation import Component_calculation
from .write_course_attainment import write_course_attainment
from .printout import printout
import os
import uuid

def driver_part1(data, Component_Details, file_path):
    wb = Workbook()
    wb.remove(wb.active)

    #prefix all the keys with the section name and replace spaces with underscore
    Component_Details = {
        f"{data['Section']}_{key.replace(' ', '_')}": value
        for key, value in Component_Details.items()
    }

    wb.create_sheet(f"{data['Section']}_Input_Details")
    ws = wb[f"{data['Section']}_Input_Details"]
    ws = input_detail(data,Component_Details,ws,conditional=True)
    ws = indirect_co_assessment(data,ws,conditional=True)
    adjust_width(ws)
    ws = CO_PO_Table(data,ws,conditional=True)

    #iterate throught Keys of Component_Details and make a worksheet for each key
    for key in Component_Details.keys():
        wb.create_sheet(key)
        ws = wb[key]
        ws.title = key
        ws = qn_co_mm_btl(data, key, Component_Details[key], ws)
        ws = studentmarks(data, key, Component_Details[key], ws)

        ws = cummulative_qn_co_mm_btl(data, key, Component_Details[key], ws)   
        ws = cummulative_studentmarks(data, key, Component_Details[key], ws)

    wb.create_sheet(f"{data['Section']}_Internal_Components")
    ws = wb[f"{data['Section']}_Internal_Components"]
    ws = Component_calculation(data,Component_Details,ws,"I")

    wb.create_sheet(f"{data['Section']}_External_Components")
    ws = wb[f"{data['Section']}_External_Components"]
    ws = Component_calculation(data,Component_Details,ws,"E")

    wb.create_sheet(f"{data['Section']}_Course_Attainment")
    ws = wb[f"{data['Section']}_Course_Attainment"]
    ws=write_course_attainment(data, Component_Details, ws)

    wb.create_sheet(f"{data['Section']}_Printout")
    ws = wb[f"{data['Section']}_Printout"]
    ws=printout(ws,data,2)

    #save workbook
    unique_id = uuid.uuid4().hex[:8]
    excel_file_name = f"{data['Section']}_{data['Batch']}_{data['Branch']}_{data['Semester']}_{data['Subject_Code']}_{unique_id}.xlsx"
    excel_file_name.replace(" ","_")
    full_path = os.path.join(file_path, excel_file_name)
    wb.save(full_path)
    return excel_file_name    

