from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl
from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl

def create_template(aw,data,COPOTable_df,final_table, result):
    
    #fill COPOTable_df with 0 for missing values
    COPOTable_df = COPOTable_df.fillna(0)


    aw.merge_cells('A1:A4')
    aw['A1'] = 'Course Outcome'
    aw['A1'].font = Font(bold=True)

    aw.merge_cells('B1:C1')
    aw['B1'] = 'Mapping with Program'
    aw['B1'].font = Font(bold=True)

    aw.merge_cells('D1:K1')
    aw['D1'] = 'Attainment % in'
    aw['D1'].font = Font(bold=True)

    aw.merge_cells('B2:B4')
    aw['B2'] = 'POs & PSOs'
    aw['B2'].font = Font(bold=True)

    aw["C2"]="Level of Mapping"
    aw["C2"].font = Font(bold=True)

    aw.merge_cells('C3:C4')
    aw['C3'] = 'Affinity'
    aw['C3'].font = Font(bold=True)

    aw.merge_cells('D2:H2')
    aw['D2'] = 'Direct'
    aw['D2'].font = Font(bold=True)

    aw.merge_cells('I2:J2')
    aw['I2'] = 'Indirect'
    aw['I2'].font = Font(bold=True)

    aw.merge_cells('K2:K3')
    aw['K2'] = 'Final Weighted CO Attainment (80% Direct + 20% Indirect)'
    aw['K2'].font = Font(bold=True)

    aw.merge_cells('D3:E3')
    aw['D3'] = 'University(SEE)'
    aw['D3'].font = Font(bold=True)

    aw.merge_cells('F3:G3')
    aw['F3'] = 'Internal(CIE)'
    aw['F3'].font = Font(bold=True)

    aw.merge_cells('H3:H4')
    aw['H3'] = 'Weighted Level of Attainment (University + IA)'
    aw['H3'].font = Font(bold=True)

    aw["D4"]="Attainment"
    aw["D4"].font = Font(bold=True)

    aw["E4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["E4"].font = Font(bold=True)

    aw["F4"]="Attainment"
    aw["F4"].font = Font(bold=True)

    aw["G4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["G4"].font = Font(bold=True)

    aw.merge_cells('I3:I4')
    aw["I3"]="Attainment"
    aw["I3"].font = Font(bold=True)

    aw.merge_cells('J3:J4')
    aw['J3']="Level Of Attainment"
    aw["J3"].font = Font(bold=True)

    aw["K4"]="Level of Attainment"
    aw["K4"].font = Font(bold=True)

    #Set column width for A to 17.22
    aw.column_dimensions['A'].width = 17.22
    aw.column_dimensions['B'].width = 9.33
    aw.column_dimensions['C'].width = 15.56
    aw.column_dimensions['D'].width = 10.33
    aw.column_dimensions['E'].width = 14.11
    aw.column_dimensions['F'].width = 10.33
    aw.column_dimensions['G'].width = 14.11
    aw.column_dimensions['H'].width = 20.67
    aw.column_dimensions['I'].width = 10.33
    aw.column_dimensions['J'].width = 18.11
    aw.column_dimensions['K'].width = 22.78
    #center align the text in the cells
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            #set color of the cells to blue
            cell.fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    #set color of the cells to green
    aw["C3"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw["K4"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')

    start=4
    interval=16

    for i in range(1, (data["Number_of_COs"]+1)):
        
        start+=1
        aw.merge_cells(start_row=start, start_column=1, end_row=start+interval, end_column=1)
        aw.cell(row=start, column=1).value = "CO"+str(i)
        aw.cell(row=start, column=1).font = Font(bold=True)
        aw.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i%2==0:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
        else:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')

        index=1
        for j in range(start, start+interval+1):
            #print out COPO mapping
            aw.cell(row=j, column=2).value = COPOTable_df.columns[index]
            aw.cell(row=j, column=3).value = COPOTable_df.iloc[i-1, index]
            if index%2==0:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
            else:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            index+=1
        
        

        for k in range(4,12):
            aw.merge_cells(start_row=start, start_column=k, end_row=start+interval, end_column=k)
            aw.cell(row=start, column=k).value = final_table.iloc[i-1, k-4]
            aw.cell(row=start, column=k).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if k%2==0:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            else:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='dce6f1', end_color='dce6f1', fill_type='solid')


            aw.cell(row=start, column=k).border = Border(left=Side(border_style='thin', color='000000'),
                                                        right=Side(border_style='thin', color='000000'),
                                                        top=Side(border_style='thin', color='000000'),
                                                        bottom=Side(border_style='thin', color='000000'))
            
    



        start=start+interval
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
    
    #find out the last row in the excel sheet
    current_row = aw.max_row+4
    aw.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=17+2)
    aw.cell(row=current_row, column=2).value = "Weighted PO/PSO Attainment Contribution"
    aw.cell(row=current_row, column=2).font = Font(bold=True)
    aw.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    

    # Print the resulting DataFrame
    # Write the header
    for c, header in enumerate(result.columns, start=2):
        aw.cell(row=current_row+1, column=c, value=header)

    # Write the DataFrame to the worksheet
    for r, row in enumerate(result.values, start=current_row+2):
        for c, value in enumerate(row, start=2):
            aw.cell(row=r, column=c, value=value)

    #create a table
    tab = Table(displayName="WeightedPO", ref="B"+str(current_row+1)+":S"+str(current_row+1+len(result)))
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    aw.add_table(tab)

    for row in aw.iter_rows(min_row=current_row, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row: 
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
    return aw