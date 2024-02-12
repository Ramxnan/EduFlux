from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries                                  #import get_column_letter from openpyxl
import time
from openpyxl import Workbook                                                         #import workbook from openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle, colors, Color, Fill, GradientFill, Font, Border, Side, Alignment, Protection
import pandas as pd
import os
import numpy as np
import uuid
from openpyxl.utils import get_column_letter
from .printout_p3 import printout
#from printout import printout_template


def driver_part3(input_dir_path, output_dir_path):
    wbwrite = Workbook()
    wbwrite.remove(wbwrite.active)
    wbwrite.create_sheet("Printouts",0)
    wbwrite.create_sheet("PO_calculations",1)
    
    wswrite_printouts=wbwrite["Printouts"]
    startrow=2
    for file in os.listdir(input_dir_path):
        if file.endswith(".xlsx") and not file.startswith("final"):
            wbread=load_workbook(input_dir_path+"\\"+file, data_only=True) 
            ws_printout=""
            for ws in wbread.sheetnames:
                if ws.endswith("Printout"):
                    ws_printout=ws
            wsread_printout=wbread[ws_printout]
            Number_of_COs=wsread_printout["B11"].value
            wswrite_printouts.merge_cells(f"D{startrow}:R{startrow}")
            wswrite_printouts[f"D{startrow}"]=file
            wswrite_printouts[f"D{startrow}"].font = Font(bold=True, size=14)
            wswrite_printouts[f"D{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
            wswrite_printouts[f"D{startrow}"].fill = PatternFill(start_color='ce875c', end_color='ce875c', fill_type='solid')
            for row in wswrite_printouts.iter_rows(min_row=startrow, max_row=startrow, min_col=4, max_col=18):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

            startrow+=1
            wswrite_printouts=printout(wswrite_printouts, Number_of_COs, startrow)
            
            min_row=1
            min_col=4
            max_row=3+Number_of_COs
            max_col=18
            for row in range(min_row, max_row+1):
                for col in range(min_col, max_col+1):
                    try:
                        wswrite_printouts.cell(row=startrow, column=col).value=wsread_printout.cell(row=row, column=col).value
                    except:
                        pass
                startrow+=1
          



            startrow+=1
            

            
    #================================================================================================
    #================================================================================================
    # #PO calculation
    wswrite_POCalculation=wbwrite["PO_calculations"]

    wswrite_POCalculation.merge_cells('A1:T1')
    wswrite_POCalculation["A1"]="PO Attainment"
    wswrite_POCalculation["A1"].font = Font(bold=True, size=18)
    wswrite_POCalculation["A1"].alignment = Alignment(horizontal='center', vertical='center')
    wswrite_POCalculation["A1"].fill = PatternFill(start_color='ffe74e', end_color='ffe74e', fill_type='solid')
    for row in wswrite_POCalculation.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    wswrite_POCalculation.merge_cells('A2:T2')
    wswrite_POCalculation["A2"]="Direct Attainment at PO level"
    wswrite_POCalculation["A2"].font = Font(bold=True, size=14)
    wswrite_POCalculation["A2"].alignment = Alignment(horizontal='center', vertical='center')
    for row in wswrite_POCalculation.iter_rows(min_row=2, max_row=2, min_col=1, max_col=20):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    wswrite_POCalculation["A3"]="S.No"
    wswrite_POCalculation["A3"].font = Font(bold=True, color="ffffff")
    wswrite_POCalculation["A3"].alignment = Alignment(horizontal='center', vertical='center')
    wswrite_POCalculation["A3"].fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation["A3"].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    wswrite_POCalculation["B3"]="Course Code"
    wswrite_POCalculation["B3"].font = Font(bold=True, color="ffffff")
    wswrite_POCalculation["B3"].alignment = Alignment(horizontal='center', vertical='center')
    wswrite_POCalculation["B3"].fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation["B3"].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    wswrite_POCalculation["C3"]="Course Name"
    wswrite_POCalculation["C3"].font = Font(bold=True, color="ffffff")
    wswrite_POCalculation["C3"].alignment = Alignment(horizontal='center', vertical='center')
    wswrite_POCalculation["C3"].fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation["C3"].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))


    po_data_row=3
    po_data_col=3
    for po in range(1, 13):
        wswrite_POCalculation.cell(row=po_data_row, column=po+po_data_col).value=f"PO{po}"
        wswrite_POCalculation.cell(row=po_data_row, column=po+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=po_data_row, column=po+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=po_data_row, column=po+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=po_data_row, column=po+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    for pso in range(1, 6):
        wswrite_POCalculation.cell(row=po_data_row, column=pso+12+po_data_col).value=f"PSO{pso}"
        wswrite_POCalculation.cell(row=po_data_row, column=pso+12+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=po_data_row, column=pso+12+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=po_data_row, column=pso+12+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=po_data_row, column=pso+12+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    columns=[]
    columns.append("Academic Year")
    columns.append("Semester")
    columns.append("Course Code")
    columns.append("Course Name")
    for i in range(1, 13):
        columns.append(f"PO{i}")
    for i in range(1, 6):
        columns.append(f"PSO{i}")
    final_po_table=pd.DataFrame(columns=columns)
    for file in os.listdir(input_dir_path):
        #file shouldnt start with final
        if file.endswith(".xlsx") and not file.startswith("final"):
            wbread=load_workbook(input_dir_path+"\\"+file, data_only=True)
            #find the name of worksheet which ends with Input_Details
            wsname_ID=""
            wsname_CA=""
            for ws in wbread.sheetnames:
                if ws.endswith("Input_Details"):
                    wsname_ID=ws
                if ws.endswith("Course_Attainment"):
                    wsname_CA=ws


            wsread_input_detials=wbread[wsname_ID]
            Number_of_COs=wsread_input_detials["B11"].value
            wsread_Course_Attainment=wbread[wsname_CA]
            row=wsread_Course_Attainment.max_row
            min_col=1
            max_col=wsread_Course_Attainment.max_column
            rowdata=[]
            for col in range(min_col, max_col+1):
                    rowdata.append(wsread_Course_Attainment.cell(row=row, column=col).value)
    #         #print(rowdata)
            rowdata_df=pd.DataFrame(rowdata).T
            rowdata_df.columns=columns
            final_po_table = pd.concat([final_po_table,rowdata_df], axis=0)
    final_po_table = final_po_table.apply(lambda x: x.astype(float) if x.dtype == int else x)
            
            
    final_po_table=final_po_table.replace(0, np.nan)
    final_po_table.reset_index(drop=True, inplace=True)

    semester_sort = {'Odd': 1, 'Even': 2}
    final_po_table['Semester code'] = final_po_table['Semester'].map(semester_sort)

    final_po_table = final_po_table.sort_values(by=['Academic Year', 'Semester code'])
    #print(final_po_table)

    dataframes_dict = {group: data.drop(['Academic Year', 'Semester', 'Semester code'], axis=1)
                   for group, data in final_po_table.groupby(['Academic Year', 'Semester'])}



    # for key, value in dataframes_dict.items():
    #     print(key)
    #     print(value)
    #     print("=====================================")

    startrow=4
    startcol=1
    sno=1
    trows=[]
    vrows=[]
    for key, value in dataframes_dict.items():
        wswrite_POCalculation.merge_cells(start_row=startrow, start_column=startcol, end_row=startrow, end_column=startcol+19)
        wswrite_POCalculation.cell(row=startrow, column=startcol).value=f"{key[0]} {key[1]}"
        wswrite_POCalculation.cell(row=startrow, column=startcol).font = Font(bold=True)
        wswrite_POCalculation.cell(row=startrow, column=startcol).alignment = Alignment(horizontal='center', vertical='center')
        wswrite_POCalculation.cell(row=startrow, column=startcol).fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
 
        for row in wswrite_POCalculation.iter_rows(min_row=startrow, max_row=startrow, min_col=startcol, max_col=startcol+19):
            for cell in row:
                cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        startrow+=1
        ridex=0
        for _ in dataframe_to_rows(value, index=False, header=False):
            vrows.append(startrow)
            cindex = 0
            for c in range(2, startcol+20):
                wswrite_POCalculation.cell(row=startrow, column=1).value=sno
                wswrite_POCalculation.cell(row=startrow, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                wswrite_POCalculation.cell(row=startrow, column=1).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
                wswrite_POCalculation.cell(row=startrow, column=c).value=value.iloc[ridex, cindex]
                wswrite_POCalculation.cell(row=startrow, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                wswrite_POCalculation.cell(row=startrow, column=c).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
                if ridex==len(value)-1 and c == startcol+20-1:
                    startrow+=1
                    trows.append(startrow)
                    wswrite_POCalculation.merge_cells(start_row=startrow, start_column=startcol, end_row=startrow, end_column=startcol+2)
                    wswrite_POCalculation.cell(row=startrow, column=startcol).value="Total"
                    for cfin in range (4, 21):
                        wswrite_POCalculation.cell(row=startrow, column=cfin).value=f"=SUM({get_column_letter(cfin)}{startrow-ridex-1}:{get_column_letter(cfin)}{startrow-1})"
                    for row in wswrite_POCalculation.iter_rows(min_row=startrow, max_row=startrow, min_col=startcol, max_col=startcol+19):
                        for cell in row:
                            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
                            cell.fill = PatternFill(start_color='fcd5b4', end_color='fcd5b4', fill_type='solid')
                            cell.alignment= Alignment(horizontal='center', vertical='center')
                            cell.font = Font(bold=True)
                    startrow+=1
                cindex+=1
            startrow+=1
            sno+=1
            ridex+=1

    #================================================================================================
    wswrite_POCalculation.merge_cells(f'A{startrow}:T{startrow}')
    wswrite_POCalculation[f"A{startrow}"]="Indirect Assessment At PO Level"
    wswrite_POCalculation[f"A{startrow}"].font = Font(bold=True, size=14)
    wswrite_POCalculation[f"A{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in wswrite_POCalculation.iter_rows(min_row=startrow, max_row=startrow, min_col=1, max_col=20):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    po_data_col=3
    startrow+=1

    wswrite_POCalculation.cell(row=startrow, column=1).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=2).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=3).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')

    for po in range(1, 13):
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).value=f"PO{po}"
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    for pso in range(1, 6):
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).value=f"PSO{pso}"
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    startrow+=1
    wswrite_POCalculation[f'A{startrow}']=sno
    sno+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=2, end_column=3)
    wswrite_POCalculation[f'B{startrow}']="Exit survey feedback"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    
    startrow+=1
    wswrite_POCalculation[f'A{startrow}']=sno
    sno+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=2, end_column=3)
    wswrite_POCalculation[f'B{startrow}']="Recruiters Feedback"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    
    startrow+=1
    wswrite_POCalculation.merge_cells(start_row=startrow, end_row=startrow, start_column=1, end_column=3)
    wswrite_POCalculation[f'A{startrow}']="Average"
    wswrite_POCalculation[f'A{startrow}'].font = Font(bold=True)

    for colind in range(4,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}']=f'=IFERROR(AVERAGE({get_column_letter(colind)}{startrow-2}:{get_column_letter(colind)}{startrow-1}),0)'

    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].fill = PatternFill(start_color='fcd5b4', end_color='fcd5b4', fill_type='solid')

    #================================================================================================
        
    startrow+=2
    wswrite_POCalculation.merge_cells(f'A{startrow}:T{startrow}')
    wswrite_POCalculation[f"A{startrow}"]='Total PO Attainment'
    wswrite_POCalculation[f"A{startrow}"].font = Font(bold=True, size=18)
    wswrite_POCalculation[f"A{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    wswrite_POCalculation[f"A{startrow}"].fill = PatternFill(start_color='95b3d7', end_color='95b3d7', fill_type='solid')
    for row in wswrite_POCalculation.iter_rows(min_row=startrow, max_row=startrow, min_col=1, max_col=20):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    startrow+=1
    wswrite_POCalculation.cell(row=startrow, column=1).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=2).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
    wswrite_POCalculation.cell(row=startrow, column=3).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')

    for po in range(1, 13):
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).value=f"PO{po}"
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=startrow, column=po+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    for pso in range(1, 6):
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).value=f"PSO{pso}"
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).font = Font(bold=True, color="ffffff")
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).fill = PatternFill(start_color='6CB266', end_color='6CB266', fill_type='solid')
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation.cell(row=startrow, column=pso+12+po_data_col).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Total Direct Assessment"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    
    for colind in range(4,21):
        formula=f'=SUM('
        for trow in trows:
            formula+=f'{get_column_letter(colind)}{trow},'
        formula=formula[:-1]
        formula+=')'
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=formula

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Total courses through PO mapped"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    print(vrows)
    for colind in range(4,21):
        # wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=mappedpo.iloc[-1, colind-4]
        formula=f'=COUNT('
        for vrow in vrows:
            formula+=f'{get_column_letter(colind)}{vrow},'
        formula=formula[:-1]
        formula+=')'
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=formula

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Average of direct Assessment"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    for colind in range(4,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=f'=IFERROR({get_column_letter(colind)}{startrow-2}/{get_column_letter(colind)}{startrow-1},0)'

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="Average of Indirect Assessment"
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    for colind in range(4,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].value=f'={get_column_letter(colind)}{startrow-7}'

    startrow+=1
    wswrite_POCalculation.merge_cells(f'A{startrow}:C{startrow}')
    wswrite_POCalculation[f'A{startrow}']="PO Attainment for the Program"
    wswrite_POCalculation[f'A{startrow}'].font = Font(bold=True, size=14)
    for colind in range(1,21):
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wswrite_POCalculation[f'{get_column_letter(colind)}{startrow}'].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    

    #set column width for second column to be 12
    wswrite_POCalculation.column_dimensions['B'].width = 12
    wswrite_POCalculation.column_dimensions['C'].width = 16


    # lastrow=[]
    # # Calculate average, excluding NaN values
    # for col in final_po_table.columns:
    #     if col == 0:
    #         lastrow.append("Average")
    #     else:
    #         # Convert the column to numeric, non-numeric values become NaN
    #         numeric_col = pd.to_numeric(final_po_table[col], errors='coerce')
    #         # Calculate the mean of the column, skipping NaN values
    #         mean_value = numeric_col.mean()
    #         lastrow.append(mean_value)

    # # Append the last row with the averages to your DataFrame
    # final_po_table.loc['Average'] = lastrow
    # final_po_table.reset_index(drop=True, inplace=True)

    # #write the final_po_table to the excel sheet from A2 to the end
    # row=2
    # for _ in dataframe_to_rows(final_po_table, index=False, header=False):
    #     for c in range(1, 19):
    #         wswrite_POCalculation.cell(row=row, column=c).value=final_po_table.iloc[row-2, c-1]
    #         wswrite_POCalculation.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    #         wswrite_POCalculation.cell(row=row, column=c).border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
    #         if row%2!=0:
    #             wswrite_POCalculation.cell(row=row, column=c).fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
            
    #     row=row+1
    # row-=1
    # for c in range(1, 19):
    #     wswrite_POCalculation.cell(row=row, column=c).fill = PatternFill(start_color='d99cac', end_color='d99cac', fill_type='solid')
    #     wswrite_POCalculation.cell(row=row, column=c).font = Font(bold=True)



    # print(final_po_t



    # ws=wbwrite["PO_calculation"]
    # final_po_table=pd.DataFrame()
    # for file in os.listdir(input_dir_path):
    #     wb=load_workbook(input_dir_path+"\\"+file, data_only=True) 
    #     ws1=wb["Course Level Attainment"]
    #     #WeightedPO
    #     potable=ws1.tables['WeightedPO']
    #     table_range=potable.ref
    #     min_col, min_row, max_col, max_row = range_boundaries(table_range)
    #     rowdata=[]
    #     for row in range(min_row, max_row+1):
    #             rowdata.append([])
    #             for col in range(min_col, max_col+1):
    #                 rowdata[-1].append(ws1.cell(row=row, column=col).value)
    #     potable = pd.DataFrame(rowdata[-1:], columns=rowdata[0])
    #     potable=potable.replace(0, np.nan)
    #     #add potable to final_po using pandas concat
    #     final_po_table=pd.concat([final_po_table,potable], axis=0)

    # #print(final_po_table)
    # lastrow=[]
    # # Calculate average, excluding NaN values
    # for col in final_po_table.columns:
    #     #print(col)
    #     if col == "COs\POs":
    #         lastrow.append("Average")
    #     else:
    #         # Convert the column to numeric, non-numeric values become NaN
    #         numeric_col = pd.to_numeric(final_po_table[col], errors='coerce')
    #         # Calculate the mean of the column, skipping NaN values
    #         mean_value = numeric_col.mean()
    #         lastrow.append(mean_value)

    # # Append the last row with the averages to your DataFrame
    # final_po_table.loc['Average'] = lastrow

    # #print(final_po_table)

    # # Add the DataFrame to the Excel sheet
    # for r in dataframe_to_rows(final_po_table, index=False, header=False):
    #     ws.append(r)
    unique_code = str(uuid.uuid4()).split("-")[0]
    file_name = f"final_{unique_code}.xlsx"
    wbwrite.save(os.path.join(output_dir_path, file_name))
    return file_name

if __name__ == "__main__":
    file1 = "C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2\\2019_19CSE345_Computer system and architecture_A_Even.xlsx"
    file2="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2\\2019_19MEE444_PCE_A_Even.xlsx"
    input_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    output_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    driver_part3(input_dir_path, output_dir_path)