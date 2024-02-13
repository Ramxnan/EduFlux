from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Protection 
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import FormulaRule
from .utils import cellstyle, cellstyle_range


def indirect_co_assessment(data,aw):
    #merge cells depending on number of POs
    aw.merge_cells(start_row=data["Number_of_COs"]+5, start_column=4, end_row=data["Number_of_COs"]+5, end_column=5)
    aw[f'D{data["Number_of_COs"]+5}']="Indirect CO Assessment"
    cellstyle_range(aw[f"D{data['Number_of_COs']+5}:E{data['Number_of_COs']+5}"], bold=True, alignment=True, border=True, fill="ffe74e")
    

    aw[f"D{data['Number_of_COs']+6}"]="COs"    
    aw[f"E{data['Number_of_COs']+6}"]="Indirect %"
    cellstyle_range(aw[f"D{data['Number_of_COs']+6}:E{data['Number_of_COs']+6}"], bold=True, alignment=True, border=True, fill="f79646")
   
    
    for nco in range(1,data["Number_of_COs"]+1):
        aw[f"D{nco+data['Number_of_COs']+6}"]=f"CO{nco}"
        cellstyle(aw[f"D{nco+data['Number_of_COs']+6}"], bold=True, alignment=True, border=True)
    
    startrow = data["Number_of_COs"]+7
    endrow = startrow + data["Number_of_COs"]-1
    cellstyle_range(aw[f"D{startrow}:E{endrow}"], alignment=True, border=True, alternate=['fcd5b4', 'fde9d9'])


    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
    for row in range(startrow, endrow+1):
        cell_coordinate = f'E{row}'
        aw.conditional_formatting.add(
            cell_coordinate,
            FormulaRule(formula=[f'ISBLANK({cell_coordinate})'], stopIfTrue=False, fill=pink_fill))
        aw.conditional_formatting.add(
            cell_coordinate,
            #greater than 100 or less than 0
            FormulaRule(formula=[f'OR({cell_coordinate}>100,{cell_coordinate}<0)'], stopIfTrue=False, fill=red_fill))
            
    

    return aw