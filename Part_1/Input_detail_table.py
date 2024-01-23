from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Protection 
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from .utils import colour_table_Input_Details

#Input details
def input_detail(data,aw):  #function to input details
    aw.merge_cells('A1:B1')
    aw['A1']="Constants"
    aw['A1'].font = Font(bold=True)
    aw['A1'].alignment = Alignment(horizontal='center', vertical='center')
    aw['A1'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw['B1'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw['A1'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    i=2
    for key, value in data.items():
        aw[f'A{i}']=key
        aw[f'A{i}'].font = Font(bold=True)
        aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))

        aw[f'B{i}']=value
        aw[f'B{i}'].font = Font(bold=True)
        aw[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))

        if i%2==0:
            aw[f'A{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
            aw[f'B{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'A{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')
            aw[f'B{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

        i+=1
    
    i+=1
    aw.merge_cells(f'A{i}:B{i}')
    aw[f'A{i}']="Variables"
    aw[f'A{i}'].font = Font(bold=True)
    aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
    aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'A{i}'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    aw['A14']="Default Threshold %"
    aw['A15']="Internal %"
    aw['A16']="External %"
    aw['B16']=f'=100-B15'
    aw['A17']="Direct %"
    aw['A18']="Indirect %"
    aw['B18']=f'=100-B17'
    aw['A19']="Target CO Attainment %"

    for i in range(14,20):
        aw[f'A{i}'].font = Font(bold=True)
        aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        if i%2==0:
            aw[f'A{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'A{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

        aw[f'B{i}'].font = Font(bold=True)
        aw[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        if i%2==0:
            aw[f'B{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'B{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    #set conditional formatting for B9 to B19 such that if its empty, it will be highlighted pink
    for i in range(14,20):
        if i!=16 and i!=18:
            aw.conditional_formatting.add(
                f'B{i}',
                FormulaRule(formula=[f'ISBLANK(B{i})'], stopIfTrue=False, fill=pink_fill)
            )

    colour_table_Input_Details(aw)
    
    return aw  
            