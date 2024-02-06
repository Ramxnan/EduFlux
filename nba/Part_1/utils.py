from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def copy_range(source_sheet, target_sheet, source_range, target_range_start_cell):
    """
    Copy values from a range of cells in one sheet to a range in another sheet.

    :param source_sheet: The sheet to copy from.
    :param target_sheet: The sheet to copy to.
    :param source_range: The range in the source sheet to copy (e.g., 'B1:G1').
    :param target_range_start_cell: The starting cell of the range in the target sheet (e.g., 'B1').
    """

    # Splitting the source range to get start and end cells
    source_start_cell, source_end_cell = source_range.split(':')
    source_start_col, source_start_row = coordinate_from_string(source_start_cell)
    source_end_col, source_end_row = coordinate_from_string(source_end_cell)

    target_start_col, target_start_row = coordinate_from_string(target_range_start_cell)

    # Loop through the source cells and copy values to the target cells
    for row in range(source_start_row, source_end_row + 1):
        for col in range(column_index_from_string(source_start_col), column_index_from_string(source_end_col) + 1):
            source_cell = source_sheet.cell(row=row, column=col)
            target_cell = target_sheet.cell(row=(row - source_start_row + target_start_row), column=(col - column_index_from_string(source_start_col) + column_index_from_string(target_start_col)))
            target_cell.value = source_cell.value


def colour_table(aw, data):
    ##write a colour code table indicating the meaning of the colours used in the sheet
    #pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    # light_red_fill = PatternFill(start_color="FF5E5E", end_color="FF5E5E", fill_type="solid")
    # yellow_fill = PatternFill(start_color="FFD9A46F", end_color="FFD9A46F", fill_type="solid")
    #pink fill means empty cell, red fill means cell value greater than expected, yellow fill means all cells values below threshold
    aw[f'A{data["Number_of_Students"]+13}']="Colour Code"
    aw[f'A{data["Number_of_Students"]+13}'].font = Font(bold=True)
    aw[f'A{data["Number_of_Students"]+14}']="Pink fill"
    aw[f'A{data["Number_of_Students"]+14}'].fill = PatternFill(start_color='D8A5B5', end_color='D8A5B5', fill_type='solid')
    aw[f'A{data["Number_of_Students"]+15}']="Red fill"
    aw[f'A{data["Number_of_Students"]+15}'].fill = PatternFill(start_color='FF5E5E', end_color='FF5E5E', fill_type='solid')
    aw[f'A{data["Number_of_Students"]+16}']="Yellow fill"
    aw[f'A{data["Number_of_Students"]+16}'].fill = PatternFill(start_color='FFD9A46F', end_color='FFD9A46F', fill_type='solid')
    aw[f'A{data["Number_of_Students"]+17}']="Blue fill"
    aw[f'A{data["Number_of_Students"]+17}'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    #set border around the cell A[Number_of_Students+13] and A[Number_of_Students+16] to thin and black

    #merge cells from B[Number_of_Students+13] to C[Number_of_Students+13]
    aw.merge_cells(f'B{data["Number_of_Students"]+13}:C{data["Number_of_Students"]+13}')
    aw.merge_cells(f'B{data["Number_of_Students"]+14}:C{data["Number_of_Students"]+14}')
    aw.merge_cells(f'B{data["Number_of_Students"]+15}:C{data["Number_of_Students"]+15}')
    aw.merge_cells(f'B{data["Number_of_Students"]+16}:C{data["Number_of_Students"]+16}')
    aw.merge_cells(f'B{data["Number_of_Students"]+17}:C{data["Number_of_Students"]+17}')

    #put what they mean in the next column\
    aw[f'B{data["Number_of_Students"]+13}']="Meaning"
    aw[f'B{data["Number_of_Students"]+13}'].font = Font(bold=True)
    aw[f'B{data["Number_of_Students"]+14}']="Empty cell"
    aw[f'B{data["Number_of_Students"]+15}']="Cell value greater than expected"
    aw[f'B{data["Number_of_Students"]+16}']="All cells values in column below threshold"
    aw[f'B{data["Number_of_Students"]+17}']="Header cell (ignore)"
  
    #set the cell colour to corresponding colour
    aw[f'B{data["Number_of_Students"]+14}'].fill = PatternFill(start_color='D8A5B5', end_color='D8A5B5', fill_type='solid')
    aw[f'B{data["Number_of_Students"]+15}'].fill = PatternFill(start_color='FF5E5E', end_color='FF5E5E', fill_type='solid')
    aw[f'B{data["Number_of_Students"]+16}'].fill = PatternFill(start_color='FFD9A46F', end_color='FFD9A46F', fill_type='solid')
    aw[f'B{data["Number_of_Students"]+17}'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')


    #set border around the cell A[Number_of_Students+13] and A[Number_of_Students+16] to thin and black
    for row in aw.iter_rows(min_row=data["Number_of_Students"]+13, max_row=data["Number_of_Students"]+17, min_col=1, max_col=3):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            
    
def colour_table_Input_Details(aw):
    #just the pink means empty cell
    row_start=aw.max_row+2
    row=row_start
    aw[f'A{row}']="Colour Code"
    aw[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    aw[f'A{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'A{row}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    row+=1
    aw[f'A{row}']="Pink fill"
    aw[f'A{row}'].fill = PatternFill(start_color='D8A5B5', end_color='D8A5B5', fill_type='solid')
    aw[f'A{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    row+=1
    aw[f'A{row}']="Red fill"
    aw[f'A{row}'].fill = PatternFill(start_color='FF5E5E', end_color='FF5E5E', fill_type='solid')
    aw[f'A{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    row=row_start
    aw[f'B{row}']="Meaning"
    aw[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    aw[f'B{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'B{row}'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    row+=1
    aw[f'B{row}']="Empty cell"
    aw[f'B{row}'].fill = PatternFill(start_color='D8A5B5', end_color='D8A5B5', fill_type='solid')
    aw[f'B{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    row+=1
    aw[f'B{row}']="Cell value greater than expected"
    aw[f'B{row}'].fill = PatternFill(start_color='FF5E5E', end_color='FF5E5E', fill_type='solid')
    aw[f'B{row}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))




def adjust_width(aw):
    #adjust width of the columns in the worksheet including the merged cells
    for col in aw.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.alignment = Alignment(horizontal='center', vertical='center')
        adjusted_width = (max_length + 2) * 1
        aw.column_dimensions[get_column_letter(column)].width = adjusted_width
