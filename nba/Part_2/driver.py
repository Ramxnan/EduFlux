from openpyxl import Workbook                                                         #import workbook from openpyxl

from Part_1.Input_Details import input_detail,CO_PO_Table,indirect_co_assessment
from Part_1.Component_values import qn_co_mm_btl,studentmarks
from Part_1.Cummulative_Component_Values import cummulative_qn_co_mm_btl,cummulative_studentmarks
from nba.Part_1.InternalExternal_Component_calculation import Component_calculation
from Part_1.write_course_attainment import write_course_attainment
from Part_1.printout import printout
from Part_1.utils import adjust_width

import os
import uuid
from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils import get_column_letter

def driver_part2(input_dir_path, output_dir_path):
    #create openpyxl workbook
    wbwrite = Workbook()
    wbwrite.remove(wbwrite.active)
    excel_files=[]
    total_students = 0
    Warnings = []


    sum_indirect_assessment = pd.DataFrame()
    sum_co_po_table = pd.DataFrame()
    sum_qn_co_mm_btl = {}
    sum_studentmarks = {}

    prev_qn_co_mm_btl_check = {}
    prev_co_po_table_check = None
    prev_Component_Details_check = None

    new_qn_co_mm_btl_check = {}
    new_co_po_table_check = None
    new_Component_Details_check = None

    for file in os.listdir(input_dir_path):
        if file.endswith(".xlsx") and not file.startswith("Sum"):
            excel_files.append(file)
    excel_files.sort()

    #All the files should have different sections
    sections = []
    for file in excel_files:
        sections.append(file[0])

    #if section is not letter, return an error
    for section in sections:
        if not section.isalpha():
            Warnings.append(f"{file} has invalid section name")
            return Warnings

    if len(sections) != len(set(sections)):
        Warnings.append("All the files should have different sections")
        return Warnings


    for file in excel_files:
        if file.endswith(".xlsx") and not file.startswith("Sum"):
            file_path = os.path.join(input_dir_path, file)
            wbread = load_workbook(file_path, data_only=True)

            alldata={}
            Component_Details = {}
            indirect_assessment = pd.DataFrame()


            input_details_title=None
            for sheet_name in wbread.sheetnames:
                if sheet_name.endswith("Input_Details"):
                    input_details_title = sheet_name

            if input_details_title is None:
                Warnings.append(f"{file} does not have Input_Details sheet")
                return Warnings

            wsread_input_details = wbread[input_details_title]
            #create a dictionary from A2 to B11
            for key, value in wsread_input_details.iter_rows(min_row=2, max_row=11, min_col=1, max_col=2, values_only=True):
                alldata[key] = value
    
            for key, value in wsread_input_details.iter_rows(min_row=14, max_row=19, min_col=1, max_col=2, values_only=True):
                alldata[key] = value

            #if any of the values in alldata is None, return an error
            if None in alldata.values():
                Warnings.append(f"{file} has missing values in Input_Details sheet")
                return Warnings

            total_students+=alldata['Number_of_Students']
            data = {key: alldata[key] for key in alldata.keys() & {'Teacher', 'Academic_year', 'Batch', 'Branch', 'Subject_Name', 'Subject_Code', 'Section', 'Semester', 'Number_of_Students', 'Number_of_COs'}}
            sum_data_all = alldata

            
                
            #extract table called Component_Details and store it in a dictionary
            table_range = wsread_input_details.tables[f'{data["Section"]}_Component_Details'].ref
            for row in wsread_input_details[table_range][1:]:
                Component_Details[row[0].value] = row[1].value

            sum_Component_Details = {f"Sum_{key[2:]}":value for key,value in Component_Details.items()}

            #Check if Component_Details is same as previous file
            new_Component_Details_check = {f"{key[2:]}":value for key,value in Component_Details.items()}
            if prev_Component_Details_check is not None:
                for key in new_Component_Details_check.keys():
                    if prev_Component_Details_check[key] != new_Component_Details_check[key]:
                        Warnings.append(f"{file} has different Component_Details")
                        return Warnings
            prev_Component_Details_check = new_Component_Details_check

            new_qn_co_mm_btl_check = {}         
            for key, qnum in Component_Details.items():
                comp_ws = wbread[key]
                
                # Define the range for QN-CO-MM-BTL details
                start_row = 2
                end_row = 7
                start_col = 3
                end_col = 3 + qnum - 1
                
                # Extract QN-CO-MM-BTL details
                data_rows = []
                for row in comp_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    data_rows.append([cell.value for cell in row])
                df_qn_co_mm_btl = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                sum_qn_co_mm_btl["Sum_"+key[2:]] = df_qn_co_mm_btl
                
                new_qn_co_mm_btl_check[key[2:]] = df_qn_co_mm_btl


                # Define the range for student marks
                start_row = 10
                end_row = 10 + alldata['Number_of_Students']
                start_col = 1
                end_col = 2 + qnum

                # Extract student marks
                data_rows = []
                for row in comp_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    data_rows.append([cell.value for cell in row])
                df_student_marks = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                if "Sum_"+key[2:] in sum_studentmarks:
                    sum_studentmarks["Sum_"+key[2:]] = pd.concat([sum_studentmarks["Sum_"+key[2:]], df_student_marks], axis=0)
                else:
                    sum_studentmarks["Sum_"+key[2:]] = df_student_marks
                 
            #Check if QN-CO-MM-BTL is same as previous file
            if prev_qn_co_mm_btl_check:
                for key in new_qn_co_mm_btl_check.keys():
                    if not prev_qn_co_mm_btl_check[key].equals(new_qn_co_mm_btl_check[key]):
                        Warnings.append(f"{file} has different QN-CO-MM-BTL Table in {key} Component")
                        return Warnings
            prev_qn_co_mm_btl_check = new_qn_co_mm_btl_check

            #get values from wsread_input_details from cell E{2+number_of_COs+4+1} to E{2+number_of_COs+4+1+number_of_COs} and concat it to sum_indirect_assessment

            start_row = 2 + data["Number_of_COs"] + 4 + 1
            end_row = 2 + data["Number_of_COs"] + 4 + data["Number_of_COs"]
            cell_range = f'E{start_row}:E{end_row}'

            values = []
            for row in wsread_input_details[cell_range]:
                for cell in row:
                    values.append(cell.value)

            indirect_assessment=pd.DataFrame(values, columns=[f"{data['Section']}"])
            sum_indirect_assessment = pd.concat([sum_indirect_assessment, indirect_assessment], axis=1)

            start_row = 3
            end_row = 3 + data["Number_of_COs"] - 1
            start_col = 5
            end_col=21
            cell_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'

            values = []
            for row in wsread_input_details[cell_range]:
                values.append([cell.value for cell in row])

            po_columns=[f"PO{i}" for i in range(1,13)]
            pso_columns=[f"PSO{i}" for i in range(1,6)]
            df_columns=po_columns+pso_columns
            print(df_columns)
            
            co_po_table=pd.DataFrame(values, columns=df_columns)
            #replace 0 with NaN
            co_po_table.replace(0, pd.NA, inplace=True)
            sum_co_po_table = co_po_table

            new_co_po_table_check = co_po_table
            if prev_co_po_table_check is not None:
                if not prev_co_po_table_check.equals(new_co_po_table_check):
                    Warnings.append(f"{file} has different CO-PO Table")
                    return Warnings
            prev_co_po_table_check = new_co_po_table_check
            
                                                         
            wbwrite.create_sheet(f"{data['Section']}_Input_Details")
            wswrite = wbwrite[f"{data['Section']}_Input_Details"]
            wswrite = input_detail(data,Component_Details,wswrite)
            wswrite = indirect_co_assessment(data,wswrite)
            adjust_width(wswrite)
            wswrite = CO_PO_Table(data,wswrite)

            for key in Component_Details.keys():
                wbwrite.create_sheet(key)
                wswrite = wbwrite[key]
                wswrite.title = key
                wswrite = qn_co_mm_btl(data, key, Component_Details[key], wswrite)
                wswrite = studentmarks(data, key, Component_Details[key], wswrite)

                wswrite = cummulative_co_mm_btl(data, key, Component_Details[key], wswrite)   
                wswrite = cummulative_studentmarks(data, key, Component_Details[key], wswrite)

            wbwrite.create_sheet(f"{data['Section']}_Internal_Components")
            wswrite = wbwrite[f"{data['Section']}_Internal_Components"]
            wswrite = Component_calculation(data,Component_Details,wswrite,"I")

            wbwrite.create_sheet(f"{data['Section']}_External_Components")
            wswrite = wbwrite[f"{data['Section']}_External_Components"]
            wswrite = Component_calculation(data,Component_Details,wswrite,"E")

            wbwrite.create_sheet(f"{data['Section']}_Course_level_Attainment")
            wswrite = wbwrite[f"{data['Section']}_Course_level_Attainment"]
            wswrite=write_course_level_attainment(data, Component_Details, wswrite)

            wbwrite.create_sheet(f"{data['Section']}_Printout")
            wswrite = wbwrite[f"{data['Section']}_Printout"]
            wswrite=printout(wswrite,data)

            #copy data from all the sheets of wbread to wbwrite
            for sheet in wbread.sheetnames:
                wsread = wbread[sheet]
                wswrite = wbwrite[sheet]
                for row in wsread.iter_rows(min_row=1, max_row=wsread.max_row, min_col=1, max_col=wsread.max_column):
                    for cell in row:
                        #if error occurs while copying the cell, skip the cell
                        try:
                            wswrite[cell.coordinate].value = cell.value
                        except:
                            pass

            
            

            

            






            wbread.close()
    #save the workbook
    
    sum_data_all['Section'] = "Sum"
    sum_data_all['Number_of_Students'] = total_students
    sum_data = {key: sum_data_all[key] for key in sum_data_all.keys() & {'Teacher', 'Academic_year', 'Batch', 'Branch', 'Subject_Name', 'Subject_Code', 'Section', 'Semester', 'Number_of_Students', 'Number_of_COs'}}


    print(total_students)
    print(sum_data_all)
    print(sum_Component_Details)
    print(sum_indirect_assessment)
    print(sum_co_po_table)
    print(sum_qn_co_mm_btl)
    print(sum_studentmarks)

    wbwrite.create_sheet(f"{sum_data['Section']}_Input_Details")
    wswrite = wbwrite[f"{sum_data['Section']}_Input_Details"]
    wswrite['B14'] = sum_data_all['Default Threshold %']
    wswrite['B15'] = sum_data_all['Internal %']
    wswrite['B17'] = sum_data_all['Direct %']
    wswrite['B19'] = sum_data_all['Target CO Attainment %']

    
            
    #add a column which is average of all the columns in sum_indirect_assessment
    sum_indirect_assessment['Avg'] = sum_indirect_assessment.mean(axis=1)
    for i in range(len(sum_indirect_assessment)):
        wswrite[f'E{2+sum_data['Number_of_COs']+4+1+i}'] = sum_indirect_assessment['Avg'][i]
    wswrite = input_detail(sum_data,sum_Component_Details,wswrite)
    
    

    wswrite = indirect_co_assessment(sum_data,wswrite)
    adjust_width(wswrite)

    start_row = 3
    end_row = 3 + data["Number_of_COs"] - 1
    start_col = 5
    end_col=21
    #paste the content of sum_co_po_table to given range in the sheet
    row=0
    col=0
    for r in range(start_row, end_row+1):
        col=0
        for c in range(start_col, end_col+1):
            wswrite.cell(row=r, column=c, value=sum_co_po_table.iloc[row,col])
            col+=1
        row+=1

    wswrite = CO_PO_Table(sum_data,wswrite)




    for key in sum_Component_Details.keys():
        #replace first letter with combined
        wbwrite.create_sheet(key)
        wswrite = wbwrite[key]
        wswrite.title = key
        wswrite = qn_co_mm_btl(sum_data, key, sum_Component_Details[key], wswrite)
        #paste the content of sum_qn_co_mm_btl to given range in the sheet
        start_row = 3
        end_row = 7
        start_col = 3
        end_col = 3 + sum_Component_Details[key] - 1
        row=0
        col=0
        try:
            for r in range(start_row, end_row+1):
                col=0
                for c in range(start_col, end_col+1):
                    wswrite.cell(row=r, column=c, value=sum_qn_co_mm_btl[key].iloc[row,col])
                    col+=1
                row+=1
        except:
            pass

        wswrite = studentmarks(sum_data, key, sum_Component_Details[key], wswrite)
        start_row = 11
        end_row = 10 + sum_data['Number_of_Students']
        start_col = 1
        end_col = 2 + sum_Component_Details[key]
        row=0
        col=0
        try:
            for r in range(start_row, end_row+1):
                col=0
                for c in range(start_col, end_col+1):
                    wswrite.cell(row=r, column=c, value=sum_studentmarks[key].iloc[row,col])
                    col+=1
                row+=1
        except:
            pass

        wswrite = cummulative_co_mm_btl(sum_data, key, sum_Component_Details[key], wswrite)
        wswrite = cummulative_studentmarks(sum_data, key, sum_Component_Details[key], wswrite)

  
    

    wbwrite.create_sheet("Sum_Internal_Components")
    wswrite = wbwrite["Sum_Internal_Components"]
    wswrite = Component_calculation(sum_data,sum_Component_Details,wswrite,"I")

    wbwrite.create_sheet("Sum_External_Components")
    wswrite = wbwrite["Sum_External_Components"]
    wswrite = Component_calculation(sum_data,sum_Component_Details,wswrite,"E")

    wbwrite.create_sheet("Sum_Course_level_Attainment")
    wswrite = wbwrite["Sum_Course_level_Attainment"]
    wswrite=write_course_level_attainment(sum_data, sum_Component_Details, wswrite)

    wbwrite.create_sheet("Sum_Printout")
    wswrite = wbwrite["Sum_Printout"]
    wswrite=printout(wswrite,sum_data)

    unique_id = str(uuid.uuid4()).split("-")[0]
    excel_file_name=f"Sum_{data['Batch']}_{data['Branch']}_{data['Semester']}_{data['Subject_Code']}_{unique_id}.xlsx"
    wbwrite.save(os.path.join(output_dir_path, excel_file_name))
    if Warnings:
        return Warnings
    else:
        return ["Files successfully merged"]

if __name__ == "__main__":
    input_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    output_dir_path="C:\\Users\\raman\\OneDrive - Amrita vishwa vidyapeetham\\ASE\\Projects\\NBA\\NBA_v3\\dev_19.1\\flux\\nba\\Part_2"
    driver_part2(input_dir_path, output_dir_path)