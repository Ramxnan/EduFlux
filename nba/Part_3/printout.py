from openpyxl.styles import Font, Alignment                                           #import font and alignment from openpyxl
from openpyxl.styles.borders import Border, Side                                #import border from openpyxl
from openpyxl.styles import PatternFill                                           #import patternfill from openpyxl


def printout_template(aw, startrow):
    aw.merge_cells(f"A{startrow}:A{startrow+2}")
    aw[f"A{startrow}"].font = Font(bold=True)
    aw[f"A{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow+2, min_col=1, max_col=1):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            

    aw.merge_cells(f"B{startrow}:B{startrow+2}")
    aw[f"B{startrow}"].font = Font(bold=True)
    aw[f"B{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow+2, min_col=2, max_col=2):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells(f"C{startrow}:D{startrow}")
    aw[f"C{startrow}"].font = Font(bold=True)
    aw[f"C{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow, min_col=3, max_col=4):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw.merge_cells(f"C{startrow+1}:D{startrow+1}")
    aw[f"C{startrow+1}"].font = Font(bold=True)
    aw[f"C{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow+1, max_row=startrow+1, min_col=3, max_col=4):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw[f"C{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"C{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f"D{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"D{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells(f"E{startrow}:F{startrow}")
    aw[f"E{startrow}"].font = Font(bold=True)
    aw[f"E{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow, min_col=5, max_col=6):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw.merge_cells(f"E{startrow}:F{startrow}")
    aw[f"E{startrow+1}"].font = Font(bold=True)
    aw[f"E{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center')

    aw[f"E{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"E{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f"F{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"F{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    

    aw.merge_cells(f"G{startrow}:H{startrow}")
    aw[f"G{startrow}"].font = Font(bold=True)
    aw[f"G{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow, min_col=7, max_col=8):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells(f"G{startrow+1}:H{startrow+1}")
    aw[f"G{startrow+1}"].font = Font(bold=True)
    aw[f"G{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow+1, max_row=startrow+1, min_col=7, max_col=8):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw[f"G{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"G{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f"H{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"H{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells(f"I{startrow}:J{startrow+1}")
    aw[f"I{startrow}"].font = Font(bold=True)
    aw[f"I{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow+1, min_col=9, max_col=10):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw[f"I{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"I{startrow}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f"J{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"J{startrow}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw.merge_cells(f"K{startrow}:L{startrow}")
    aw[f"K{startrow}"].font = Font(bold=True)
    aw[f"K{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=startrow, max_row=startrow, min_col=11, max_col=12):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
            
    aw.merge_cells(f"K{startrow+1}:L{startrow+1}")
    aw[f"K{startrow+1}"].font = Font(bold=True)
    aw[f"K{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in aw.iter_rows(min_row=startrow+1, max_row=startrow+1, min_col=11, max_col=12):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw[f"K{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"K{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw[f"L{startrow+2}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"L{startrow+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))

    aw[f"M{startrow}"].font = Font(bold=True)
    aw[f"M{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"M{startrow}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
            
    aw[f"M{startrow+1}"].font = Font(bold=True)
    aw[f"M{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"M{startrow+1}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    aw[f"N{startrow}"].font = Font(bold=True)
    aw[f"N{startrow}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"N{startrow}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    aw[f"N{startrow+1}"].font = Font(bold=True)
    aw[f"N{startrow+1}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"N{startrow+1}"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    
    
    aw.column_dimensions['A'].width = 8.43
    aw.column_dimensions['B'].width = 8.43
    aw.column_dimensions['C'].width = 12
    aw.column_dimensions['D'].width = 12
    aw.column_dimensions['E'].width = 12
    aw.column_dimensions['F'].width = 12
    aw.column_dimensions['G'].width = 12
    aw.column_dimensions['H'].width = 12
    aw.column_dimensions['I'].width = 12
    aw.column_dimensions['J'].width = 8.43
    aw.column_dimensions['K'].width = 20
    aw.column_dimensions['L'].width = 8.43
    aw.column_dimensions['M'].width = 8.43
    aw.column_dimensions['N'].width = 10

    for row in aw.iter_cols(min_row=startrow+2, max_row=startrow+2, min_col=3, max_col=14):
        for cell in row:
            cell.fill = PatternFill(start_color='8db4e2', end_color='8db4e2', fill_type='solid')
