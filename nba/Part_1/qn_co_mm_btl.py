from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Font                                       #import font and alignment from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Protection


#qn_co_mm_btl table
def qn_co_mm_btl(data,key, Component_details,aw):  #function to create qn_co_mm_btl table
    aw.merge_cells(f'B1:{get_column_letter(Component_details+2)}1')
    aw[f'B1']=key
    
    aw['B2']="Question"
    aw['B3']="Max Marks"
    aw['B4']="Threshold"
    aw['B5']="CO"
    aw['B6']="Final CO"
    aw['B7']="BTL"
    

    for qno in range(1,Component_details+1):
        aw[get_column_letter(qno+2)+'2']=f"Q{qno}"        
        aw[get_column_letter(qno+2) + '6'].value = f'=CONCATENATE("{data["Subject_Code"]+"_CO"}", {get_column_letter(qno+2)}5)'
        aw[get_column_letter(qno+2)+'4']=f'=Input_Details!B14/100*{get_column_letter(qno+2)}3'
        

    light_red_fill = PatternFill(start_color="FF5E5E", end_color="FF5E5E", fill_type="solid")
    green_fill=PatternFill(start_color="5e9955", end_color="5e9955", fill_type='solid')                                                                                                                     #add table to workshee
    # Lock all cells by default
    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    aw.protection.sheet = True

    for qno in range(1, Component_details + 1):
        threshold_cell = f'{get_column_letter(qno+2)}4'
        max_marks_cell = f'{get_column_letter(qno+2)}3'
        co_cell = f'{get_column_letter(qno+2)}5'
        btl_cell = f'{get_column_letter(qno+2)}7'

        #set conditional formatting for max marks cell such that if its more than 100, it will be highlighted red
        rule = CellIsRule(operator='greaterThan', formula=['100'], stopIfTrue=True, fill=light_red_fill)
        aw.conditional_formatting.add(max_marks_cell, rule)
        aw[max_marks_cell].protection = Protection(locked=False)           
        
        # Conditional Formatting Rule
        rule = CellIsRule(operator='greaterThan', formula=[max_marks_cell], stopIfTrue=True, fill=light_red_fill)
        aw.conditional_formatting.add(threshold_cell, rule)
        aw[threshold_cell].protection = Protection(locked=False)

        # Conditional Formatting Rule for co cell such that if its more than data["Number_of_COs"], it will be highlighted red
        rule = CellIsRule(operator='greaterThan', formula=[data["Number_of_COs"]], stopIfTrue=True, fill=light_red_fill)
        aw.conditional_formatting.add(co_cell, rule)
        aw[co_cell].protection = Protection(locked=False)

        # Conditional Formatting Rule for btl cell such that if its more than 100, it will be highlighted red
        rule = CellIsRule(operator='greaterThan', formula=['100'], stopIfTrue=True, fill=light_red_fill)
        aw.conditional_formatting.add(btl_cell, rule)
        aw[btl_cell].protection = Protection(locked=False)


    # do the same highlight empty cells for qn_co_mm_btl table
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    for qno in range(1, Component_details + 1):
        column_letter = get_column_letter(qno + 2)
        data_range = f"{column_letter}3:{column_letter}7"
        aw.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'ISBLANK({column_letter}3)'], stopIfTrue=False, fill=pink_fill)
        )

    #set cell B2 to B7 to colour c0504d and white font
    for row in aw.iter_rows(min_row=2, max_row=7, min_col=2, max_col=2):
        for cell in row:
            cell.fill = PatternFill(fgColor="4bacc6", fill_type = "solid")
            cell.font = Font(color="FFFFFF")
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
            cell.font = Font(color="FFFFFF")
            cell.font = Font(bold=True)

    table_range = f"C2:{get_column_letter(Component_details + 2)}7"
    tab = Table(displayName=f"qn_co_mm_btl_{key}", ref=table_range)
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style
    aw.add_table(tab)

    return aw