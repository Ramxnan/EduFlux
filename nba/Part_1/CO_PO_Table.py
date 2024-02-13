from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Protection 
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import FormulaRule
from .utils import cellstyle, cellstyle_range


def CO_PO_Table(data,aw):
    #merge cells depending on number of POs
    aw.merge_cells(start_row=1, start_column=4, end_row=1, end_column=12+5+1+3)
    aw['D1']="CO-PO Mapping"
    cellstyle(aw['D1'], bold=True, alignment=True, border=True, fill="ffe74e")
   
    aw["D2"]="COs\\POs"
    cellstyle(aw["D2"], bold=True, alignment=True, border=True, fill="9bbb59")


    for nco in range(1,data["Number_of_COs"]+1):
        aw[f"D{nco+2}"]=f"CO{nco}"
        cellstyle(aw[f"D{nco+2}"], bold=True, alignment=True, border=True)

    for popso in range(1,12+5+1):
        if popso<=12:
            aw[f"{get_column_letter(popso+4)}2"]=f"PO{popso}   "
        else:
            aw[f"{get_column_letter(popso+4)}2"]=f"PSO{popso-12}"
        cellstyle(aw[f"{get_column_letter(popso+4)}2"], bold=True, alignment=True, border=True, fill="9bbb59")

    cellstyle_range(aw[f"D3:U{2+data['Number_of_COs']}"],border=True, alternate=['ebf1de','ffffff'])
            
        
    #for columns 4 to 4+12+5 set width to 13
    for col in range(4,4+12+5):
        aw.column_dimensions[f"{get_column_letter(col)}"].width = 13

   

            
    #set conditional formatting for empty cells
        
    pink_fill = PatternFill(start_color="D8A5B5", end_color="D8A5B5", fill_type="solid")
    red_fill = PatternFill(start_color="ff5e5e", end_color="ff5e5e", fill_type="solid")
    for nco in range(1,data["Number_of_COs"]+1):
        for popso in range(1,12+5+1):
            aw.conditional_formatting.add(
                f"{get_column_letter(popso+4)}{nco+2}",
                FormulaRule(formula=[f'ISBLANK({get_column_letter(popso+4)}{nco+2})'], stopIfTrue=False, fill=pink_fill))
            aw.conditional_formatting.add(
                f"{get_column_letter(popso+4)}{nco+2}",
                #greater than 100 or less than 0
                FormulaRule(formula=[f'OR({get_column_letter(popso+4)}{nco+2}>3,{get_column_letter(popso+4)}{nco+2}<0)'], stopIfTrue=False, fill=red_fill))
         
    return aw