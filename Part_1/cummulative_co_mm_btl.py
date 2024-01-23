from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Font                                       #import font and alignment from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import CellIsRule, FormulaRule

def cummulative_co_mm_btl(data, component_name, component_details, aw):
    #from number of questions + 3 start printing the numnber of cos
    user_input_end_col = get_column_letter(component_details + 2)
    user_input_range = f'C5:{user_input_end_col}5'

    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(component_details+3+cno)}2'] = f'CO{cno}'
       
        co_name = f'{data["Subject_Code"]}_CO{cno}'  # Replace with the actual CO name you're checking
        sum_range_start_maxmarks = 'C3'
        sum_range_end_maxmarks = f'{get_column_letter(component_details + 2)}3'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(component_details + 2)}6'

        aw[f'{get_column_letter(component_details+3+cno)}3'] = f'=SUMIFS({sum_range_start_maxmarks}:{sum_range_end_maxmarks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'

        sum_range_start_threshold = 'C4'
        sum_range_end_threshold = f'{get_column_letter(component_details + 2)}4'
        aw[f'{get_column_letter(component_details+3+cno)}4'] = f'=SUMIFS({sum_range_start_threshold}:{sum_range_end_threshold}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'
        

    #make it a table
    table_range = f"{get_column_letter(component_details + 4)}2:{get_column_letter(component_details + 3 + data['Number_of_COs'])}4"
    tab = Table(displayName=f"cummulative_co_mm_btl_{component_name}", ref=table_range)
    style = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style
    aw.add_table(tab)

    return aw
    