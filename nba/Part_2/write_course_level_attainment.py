import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection

def write_course_level_attainment(data,Component_Details,aw):

    aw.merge_cells('A1:A4')
    aw['A1'] = 'Course Outcome'
    aw['A1'].font = Font(bold=True)

    aw.merge_cells('B1:C1')
    aw['B1'] = 'Mapping with Program'
    aw['B1'].font = Font(bold=True)

    aw.merge_cells('D1:K1')
    aw['D1'] = 'Attainment % in'
    aw['D1'].font = Font(bold=True)

    aw.merge_cells('B2:B4')
    aw['B2'] = 'POs & PSOs'
    aw['B2'].font = Font(bold=True)

    aw["C2"]="Level of Mapping"
    aw["C2"].font = Font(bold=True)

    aw.merge_cells('C3:C4')
    aw['C3'] = 'Affinity'
    aw['C3'].font = Font(bold=True)

    aw.merge_cells('D2:H2')
    aw['D2'] = 'Direct'
    aw['D2'].font = Font(bold=True)

    aw.merge_cells('I2:J2')
    aw['I2'] = 'Indirect'
    aw['I2'].font = Font(bold=True)

    aw.merge_cells('K2:K3')
    aw['K2'] = 'Final Weighted CO Attainment (80% Direct + 20% Indirect)'
    aw['K2'].font = Font(bold=True)

    aw.merge_cells('D3:E3')
    aw['D3'] = 'University(SEE)'
    aw['D3'].font = Font(bold=True)

    aw.merge_cells('F3:G3')
    aw['F3'] = 'Internal(CIE)'
    aw['F3'].font = Font(bold=True)

    aw.merge_cells('H3:H4')
    aw['H3'] = 'Weighted Level of Attainment (University + IA)'
    aw['H3'].font = Font(bold=True)

    aw["D4"]="Attainment"
    aw["D4"].font = Font(bold=True)

    aw["E4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["E4"].font = Font(bold=True)

    aw["F4"]="Attainment"
    aw["F4"].font = Font(bold=True)

    aw["G4"]="Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)"
    aw["G4"].font = Font(bold=True)

    aw.merge_cells('I3:I4')
    aw["I3"]="Attainment"
    aw["I3"].font = Font(bold=True)

    aw.merge_cells('J3:J4')
    aw['J3']="Level Of Attainment"
    aw["J3"].font = Font(bold=True)

    aw["K4"]="Level of Attainment"
    aw["K4"].font = Font(bold=True)

    #Set column width for A to 17.22
    aw.column_dimensions['A'].width = 17.22
    aw.column_dimensions['B'].width = 9.33
    aw.column_dimensions['C'].width = 15.56
    aw.column_dimensions['D'].width = 10.33
    aw.column_dimensions['E'].width = 14.11
    aw.column_dimensions['F'].width = 10.33
    aw.column_dimensions['G'].width = 14.11
    aw.column_dimensions['H'].width = 20.67
    aw.column_dimensions['I'].width = 10.33
    aw.column_dimensions['J'].width = 18.11
    aw.column_dimensions['K'].width = 22.78
    #center align the text in the cells
    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            #set color of the cells to blue
            cell.fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    #set color of the cells to green
    aw["C3"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw["K4"].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')

    start=4
    interval=16
    rowindex=1
    for i in range(1, (data["Number_of_COs"]+1)):
        start+=1
        aw.merge_cells(start_row=start, start_column=1, end_row=start+interval, end_column=1)
        aw.cell(row=start, column=1).value = "CO"+str(i)
        aw.cell(row=start, column=1).font = Font(bold=True)
        aw.cell(row=start, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i%2==0:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
        else:
            aw.cell(row=start, column=1).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')


        index=1
        for j in range(start, start+interval+1):
            #print out COPO mapping
            aw.cell(row=j, column=2).value =f'={data["Section"]}_Input_Details!{get_column_letter(index+4)}2'
            aw.cell(row=j, column=2).alignment = Alignment(horizontal='center', vertical='center')
            aw.cell(row=j, column=2).border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
            aw.cell(row=j, column=3).value = f'={data["Section"]}_Input_Details!{get_column_letter(index+4)}{i+2}'
            aw.cell(row=j, column=3).alignment = Alignment(horizontal='center', vertical='center')
            aw.cell(row=j, column=3).border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
            if index%2==0:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
            else:
                aw.cell(row=j, column=2).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
                aw.cell(row=j, column=3).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            index+=1

        for k in range(4,12):
            aw.merge_cells(start_row=start, start_column=k, end_row=start+interval, end_column=k)
            #aw.cell(row=start, column=k).value = final_table.iloc[i-1, k-4]
            aw.cell(row=start, column=k).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if k%2==0:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            else:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='dce6f1', end_color='dce6f1', fill_type='solid')
    
        internal_components_num=0
        external_components_num=0
        for component_name in Component_Details.keys():
            if component_name[-1]=="I":
                internal_components_num+=1
            else:
                external_components_num+=1
        col=(data["Number_of_COs"]*internal_components_num) + (1*internal_components_num) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=4).value=f'=IFERROR({data["Section"]}_Internal_Components!{get_column_letter(col)}{row}, 0)'

        aw.cell(row=start, column=5).value=f'=IF(AND({get_column_letter(4)}{start}>=0,{get_column_letter(4)}{start}<40),1,IF(AND({get_column_letter(4)}{start}>=40,{get_column_letter(4)}{start}<60),2,IF(AND({get_column_letter(4)}{start}>=60,{get_column_letter(4)}{start}<=100),3,"0")))'

        col=(data["Number_of_COs"]*external_components_num) + (1*external_components_num) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=6).value=f'=IFERROR({data["Section"]}_External_Components!{get_column_letter(col)}{row}, 0)'

        aw.cell(row=start, column=7).value=f'=IF(AND({get_column_letter(6)}{start}>=0,{get_column_letter(6)}{start}<40),1,IF(AND({get_column_letter(6)}{start}>=40,{get_column_letter(6)}{start}<60),2,IF(AND({get_column_letter(6)}{start}>=60,{get_column_letter(6)}{start}<=100),3,"0")))'

        column_5_cell = f'{get_column_letter(5)}{start}'
        column_7_cell = f'{get_column_letter(7)}{start}'
        calculation = f'{column_5_cell}*({data["Section"]}_Input_Details!B16/100)+{column_7_cell}*{data["Section"]}_Input_Details!B15/100'

        formula = f'={calculation}'
        aw.cell(row=start, column=8).value = formula

        aw.cell(row=start, column=9).value=f'=IF({data["Section"]}_Input_Details!E{2+data["Number_of_COs"]+4+rowindex}>0,{data["Section"]}_Input_Details!E{2+data["Number_of_COs"]+4+rowindex},"0")'

        aw.cell(row=start, column=10).value=f'=IF(AND({get_column_letter(9)}{start}>=0,{get_column_letter(9)}{start}<40),1,IF(AND({get_column_letter(9)}{start}>=40,{get_column_letter(9)}{start}<60),2,IF(AND({get_column_letter(9)}{start}>=60,{get_column_letter(9)}{start}<=100),3,"0")))'

        aw.cell(row=start, column=11).value=f'=({get_column_letter(8)}{start}*({data["Section"]}_Input_Details!B17/100))+({get_column_letter(10)}{start}*({data["Section"]}_Input_Details!B18/100))'
        
        rowindex+=1
        start=start+interval

    for row in aw.iter_rows(min_row=1, max_row=aw.max_row, min_col=1, max_col=aw.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            

    #================================================================================================================================================================
    current_row = aw.max_row+4
    current_col = 2
    aw.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=17+2)
    aw.cell(row=current_row, column=2).value = "Weighted PO/PSO Attainment Contribution"
    aw.cell(row=current_row, column=2).font = Font(bold=True)
    aw.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=2).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    

    current_row+=1
    aw.cell(row=current_row, column=2).value = "COs\\POs"
    aw.cell(row=current_row, column=2).font = Font(bold=True, color="FFFFFF")
    aw.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=2).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row, column=2).fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    for po in range(1,12+1):
        aw[f"{get_column_letter(po+2)}{current_row}"]=f"PO{po}"
        aw[f"{get_column_letter(po+2)}{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"{get_column_letter(po+2)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(po+2)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(po+2)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        
    
    for pso in range(1,6):
        aw[f"{get_column_letter(12+2+pso)}{current_row}"]=f"PSO{pso}"
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        

    current_row+=1
    current_col=3

    start=4
    interval=17


    for co in range(1,data["Number_of_COs"]+1):
        aw[f"B{current_row}"]=f"CO{co}"
        aw[f"B{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"B{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"B{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

        for po in range(1,12+1):
            aw[f"{get_column_letter(po+2)}{current_row}"]=f'=C{start+po}*K{start+1}'
            aw[f"{get_column_letter(po+2)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            aw[f"{get_column_letter(po+2)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
            
            

        for pso in range(1,6):
            aw[f"{get_column_letter(12+2+pso)}{current_row}"]=f'=C{start+12+pso}*K{start+1}'
            aw[f"{get_column_letter(12+2+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            aw[f"{get_column_letter(12+2+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))

        current_row+=1
        start+=interval

    aw.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=17+2)
    aw.cell(row=current_row, column=2).value = "Final Ratio"
    aw.cell(row=current_row, column=2).font = Font(bold=True)
    aw.cell(row=current_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=2).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')

    current_row+=1
    current_col=2
    aw[f"{get_column_letter(current_col)}{current_row}"]=f"{data['Subject_Code']}"
    aw[f"{get_column_letter(current_col)}{current_row}"].font = Font(bold=True, color="FFFFFF")
    aw[f"{get_column_letter(current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"{get_column_letter(current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
    aw[f"{get_column_letter(current_col)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    for po in range(1,12+1):
        main_formula = f'SUM({get_column_letter(po+2)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(po+2)}{current_row-2})/(SUM({data["Section"]}_Input_Details!{get_column_letter(4+po)}3:{data["Section"]}_Input_Details!{get_column_letter(4+po)}{data["Number_of_COs"]+2}))'
        complete_formula = f'=IF(AND(SUM({get_column_letter(po+2)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(po+2)}{current_row-2})>0, SUM({data["Section"]}_Input_Details!{get_column_letter(4+po)}3:{data["Section"]}_Input_Details!{get_column_letter(4+po)}{data["Number_of_COs"]+2})>0), {main_formula}, 0)'
        aw[f"{get_column_letter(po+2)}{current_row}"] = complete_formula
        aw[f"{get_column_letter(po+2)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(po+2)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    for pso in range(1,6):
        main_formula = f'SUM({get_column_letter(12+2+pso)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(12+2+pso)}{current_row-2})/(SUM({data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}3:{data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}{data["Number_of_COs"]+2}))'
        complete_formula = f'=IF(AND(SUM({get_column_letter(12+2+pso)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(12+2+pso)}{current_row-2})>0, SUM({data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}3:{data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}{data["Number_of_COs"]+2})>0), {main_formula}, 0)'
        aw[f"{get_column_letter(12+2+pso)}{current_row}"] = complete_formula
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(12+2+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    current_row+=1


    for row in aw.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=True)
    aw.protection.sheet = True

    

    return aw