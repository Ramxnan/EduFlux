from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Font                                       #import font and alignment from openpyxl
from openpyxl.utils import get_column_letter                                  #import get_column_letter from openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo          #import table from openpyxl
from openpyxl.styles import Color, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

def cummulative_co_mm_btl(data, component_name, component_details, aw):
    #from number of questions + 3 start printing the numnber of cos
    user_input_end_col = get_column_letter(component_details + 2)
    user_input_range = f'C5:{user_input_end_col}5'

    for cno in range(1, data["Number_of_COs"]+1):
        aw[f'{get_column_letter(component_details+3+cno)}2'] = f'CO{cno}'
        aw[f'{get_column_letter(component_details+3+cno)}2'].font = Font(bold=True, color='FFFFFF')
        aw[f'{get_column_letter(component_details+3+cno)}2'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        
       
        co_name = f'{data["Subject_Code"]}_CO{cno}'  # Replace with the actual CO name you're checking
        sum_range_start_maxmarks = 'C3'
        sum_range_end_maxmarks = f'{get_column_letter(component_details + 2)}3'
        criteria_range_start = 'C6'
        criteria_range_end = f'{get_column_letter(component_details + 2)}6'

        aw[f'{get_column_letter(component_details+3+cno)}3'] = f'=SUMIFS({sum_range_start_maxmarks}:{sum_range_end_maxmarks}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'

        sum_range_start_threshold = 'C4'
        sum_range_end_threshold = f'{get_column_letter(component_details + 2)}4'
        aw[f'{get_column_letter(component_details+3+cno)}4'] = f'=SUMIFS({sum_range_start_threshold}:{sum_range_end_threshold}, {criteria_range_start}:{criteria_range_end}, "{co_name}")'
        
    for cno in range(1, data["Number_of_COs"]+1):
        #apply border and center alignment to the cells
        aw[f'{get_column_letter(component_details+3+cno)}2'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
        aw[f'{get_column_letter(component_details+3+cno)}2'].alignment = Alignment(horizontal='center', vertical='center')

        aw[f'{get_column_letter(component_details+3+cno)}3'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
        aw[f'{get_column_letter(component_details+3+cno)}3'].alignment = Alignment(horizontal='center', vertical='center')

        aw[f'{get_column_letter(component_details+3+cno)}4'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
        aw[f'{get_column_letter(component_details+3+cno)}4'].alignment = Alignment(horizontal='center', vertical='center')



    return aw
    