import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection

def write_course_attainment(data,Component_Details,aw):
    #============================================================================================================

    aw.merge_cells('A1:B1')
    aw['A1']="Constants"
    aw['A1'].font = Font(bold=True)
    aw['A1'].alignment = Alignment(horizontal='center', vertical='center')
    aw['A1'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw['B1'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw['A1'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    i=2
    for key, value in data.items():
        aw[f'A{i}']=key
        aw[f'A{i}'].font = Font(bold=True)
        aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))

        aw[f'B{i}']=value
        aw[f'B{i}'].font = Font(bold=True)
        aw[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))

        if i%2==0:
            aw[f'A{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
            aw[f'B{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'A{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')
            aw[f'B{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

        i+=1
    
    i+=1
    aw.merge_cells(f'A{i}:B{i}')
    aw[f'A{i}']="Variables"
    aw[f'A{i}'].font = Font(bold=True)
    aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
    aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw[f'A{i}'].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    aw['A14']="Default Threshold %"
    aw['B14']=f'={data["Section"]}_Input_Details!B14'
    aw['A15']="Internal %"
    aw['B15']=f'={data["Section"]}_Input_Details!B15'
    aw['A16']="External %"
    aw['B16']=f'={data["Section"]}_Input_Details!B16'
    aw['A17']="Direct %"
    aw['B17']=f'={data["Section"]}_Input_Details!B17'
    aw['A18']="Indirect %"
    aw['B18']=f'={data["Section"]}_Input_Details!B18'
    aw['A19']="Target CO Attainment %"
    aw['B19']=f'={data["Section"]}_Input_Details!B19'

    for i in range(14,20):
        aw[f'A{i}'].font = Font(bold=True)
        aw[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'A{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        if i%2==0:
            aw[f'A{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'A{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

        aw[f'B{i}'].font = Font(bold=True)
        aw[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
        aw[f'B{i}'].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        if i%2==0:
            aw[f'B{i}'].fill = PatternFill(start_color='b7dee8', end_color='b7dee8', fill_type='solid')
        else:
            aw[f'B{i}'].fill = PatternFill(start_color='daeef3', end_color='daeef3', fill_type='solid')

    aw.column_dimensions['A'].width = 23
    aw.column_dimensions['B'].width = 12
    aw.column_dimensions['C'].width = 23

    #============================================================================================================
    #============================================================================================================

    #merge cells depending on number of POs
    aw.merge_cells(start_row=1, start_column=4, end_row=1, end_column=12+5+1+3)
    aw['D1']="CO-PO Mapping"
    aw['D1'].font = Font(bold=True)
    aw['D1'].alignment = Alignment(horizontal='center', vertical='center')
    aw['D1'].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw['D1'].fill = PatternFill(start_color='ffe74e', end_color='ffe74e', fill_type='solid')

    aw["D2"]="COs\\POs"
    aw["D2"].font = Font(bold=True)
    aw["D2"].alignment = Alignment(horizontal='center', vertical='center')
    aw["D2"].border = Border(top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'),
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'))
    aw["D2"].fill = PatternFill(start_color='9bbb59', end_color='9bbb59', fill_type='solid')


    for co in range(1,data["Number_of_COs"]+1):
        aw[f"D{co+2}"]=f"CO{co}"
        aw[f"D{co+2}"].font = Font(bold=True)
        aw[f"D{co+2}"].alignment = Alignment(horizontal='center', vertical='center')
        for po in range(1,12+1):
            aw[f"{get_column_letter(po+4)}{co+2}"]=f'={data["Section"]}_Input_Details!{get_column_letter(po+4)}{co+2}'
            aw[f"{get_column_letter(po+4)}{co+2}"].alignment = Alignment(horizontal='center', vertical='center')
        for pso in range(1,6):
            aw[f"{get_column_letter(12+4+pso)}{co+2}"]=f'={data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}{co+2}'
            aw[f"{get_column_letter(12+4+pso)}{co+2}"].alignment = Alignment(horizontal='center', vertical='center')

    for po in range(1,12+1):
        aw[f"{get_column_letter(po+4)}2"]=f"PO{po}   "
        aw[f"{get_column_letter(po+4)}2"].font = Font(bold=True)
        aw[f"{get_column_letter(po+4)}2"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(po+4)}2"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(po+4)}2"].fill = PatternFill(start_color='9bbb59', end_color='9bbb59', fill_type='solid')
    for pso in range(1,6):
        aw[f"{get_column_letter(12+4+pso)}2"]=f"PSO{pso}"
        aw[f"{get_column_letter(12+4+pso)}2"].font = Font(bold=True)
        aw[f"{get_column_letter(12+4+pso)}2"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(12+4+pso)}2"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(12+4+pso)}2"].fill = PatternFill(start_color='9bbb59', end_color='9bbb59', fill_type='solid')
    
    for co in range(1,data["Number_of_COs"]+1):
        for po in range(12+1):
            aw[f"{get_column_letter(po+4)}{co+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
            if co%2==0:
                aw[f"{get_column_letter(po+4)}{co+2}"].fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
            else:
                aw[f"{get_column_letter(po+4)}{co+2}"].fill = PatternFill(start_color='ebf1de', end_color='ebf1de', fill_type='solid')

        for pso in range(1,6):
            aw[f"{get_column_letter(12+4+pso)}{co+2}"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
            if co%2==0:
                aw[f"{get_column_letter(12+4+pso)}{co+2}"].fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
            else:
                aw[f"{get_column_letter(12+4+pso)}{co+2}"].fill = PatternFill(start_color='ebf1de', end_color='ebf1de', fill_type='solid')
  

    #============================================================================================================
    #============================================================================================================
    #merge cells depending on number of POs
    aw.merge_cells(start_row=data["Number_of_COs"]+5, start_column=4, end_row=data["Number_of_COs"]+5, end_column=5)
    aw[f'D{data["Number_of_COs"]+5}']="Indirect CO Assessment"
    aw[f'D{data["Number_of_COs"]+5}'].font = Font(bold=True)
    aw[f'D{data["Number_of_COs"]+5}'].alignment = Alignment(horizontal='center', vertical='center')
    aw[f'D{data["Number_of_COs"]+5}'].fill = PatternFill(start_color="ffe74e", end_color="ffe74e", fill_type="solid")
    for row in aw.iter_rows(min_row=data["Number_of_COs"]+5, max_row=data["Number_of_COs"]+5, min_col=4, max_col=5):
        for cell in row:
            cell.border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))

    aw[f"D{data['Number_of_COs']+6}"]="COs"
    aw[f"D{data['Number_of_COs']+6}"].font = Font(bold=True)
    aw[f"D{data['Number_of_COs']+6}"].fill = PatternFill(start_color='f79646', end_color='f79646', fill_type='solid')
    aw[f"D{data['Number_of_COs']+6}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
    aw[f"D{data['Number_of_COs']+6}"].alignment = Alignment(horizontal='center', vertical='center')
    

    for i in range(1,data["Number_of_COs"]+1):
        aw[f"D{i+data['Number_of_COs']+6}"]=f"CO{i}"
        aw[f"D{i+data['Number_of_COs']+6}"].font = Font(bold=True)
        aw[f"D{i+data['Number_of_COs']+6}"].alignment = Alignment(horizontal='center', vertical='center')
        #border
        aw[f"D{i+data['Number_of_COs']+6}"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        aw[f"E{i+data['Number_of_COs']+6}"].border = Border(top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'),
                                    left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'))
        aw[f"E{i+data['Number_of_COs']+6}"]=f'={data["Section"]}_Input_Details!E{i+data["Number_of_COs"]+6}'
        aw[f"E{i+data['Number_of_COs']+6}"].alignment = Alignment(horizontal='center', vertical='center')

        if i%2==0:
            aw[f"D{i+data['Number_of_COs']+6}"].fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
            aw[f"E{i+data['Number_of_COs']+6}"].fill = PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
        else:
            aw[f"D{i+data['Number_of_COs']+6}"].fill = PatternFill(start_color='fcd5b4' , end_color='fcd5b4', fill_type='solid')
            aw[f"E{i+data['Number_of_COs']+6}"].fill = PatternFill(start_color='fcd5b4' , end_color='fcd5b4', fill_type='solid')

    aw[f"E{data['Number_of_COs']+6}"]="Indirect %"
    aw[f"E{data['Number_of_COs']+6}"].font = Font(bold=True)
    aw[f"E{data['Number_of_COs']+6}"].fill = PatternFill(start_color='f79646', end_color='f79646', fill_type='solid')
    aw[f"E{data['Number_of_COs']+6}"].border = Border(top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'))
                
    #============================================================================================================
    #============================================================================================================

    #merge cells col 4 to 16 and rows 8 to 10
    start_col_ca = 4

    aw.merge_cells(start_row=data["Number_of_COs"]+data["Number_of_COs"]+8, start_column=start_col_ca, end_row=data["Number_of_COs"]+data["Number_of_COs"]+10, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca)}{data["Number_of_COs"]+data["Number_of_COs"]+8}'].fill = PatternFill(start_color='ffe74e', end_color='ffe74e', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca)}{data["Number_of_COs"]+data["Number_of_COs"]+8}'] = 'Course Attainment'
    aw[f'{get_column_letter(start_col_ca)}{data["Number_of_COs"]+data["Number_of_COs"]+8}'].font = Font(bold=True, size=20)
    aw[f'{get_column_letter(start_col_ca)}{data["Number_of_COs"]+data["Number_of_COs"]+8}'].alignment = Alignment(horizontal='center', vertical='center')
    for row in aw.iter_rows(min_row=data["Number_of_COs"]+data["Number_of_COs"]+8, max_row=data["Number_of_COs"]+data["Number_of_COs"]+10, min_col=start_col_ca, max_col=start_col_ca+12):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            cell.fill = PatternFill(start_color='ffe74e', end_color='ffe74e', fill_type='solid')



    start_row_ca = data["Number_of_COs"]+8+data["Number_of_COs"]+3

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca, end_row=start_row_ca+3, end_column=start_col_ca)
    aw[f'{get_column_letter(start_col_ca)}{start_row_ca}'] = 'Course Outcome'
    aw[f'{get_column_letter(start_col_ca)}{start_row_ca}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca+1, end_row=start_row_ca, end_column=start_col_ca+2)
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca}'] = 'Mapping with Program'
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+1, end_row=start_row_ca+3, end_column=start_col_ca+1)
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca+1}'] = 'POs & PSOs'
    aw[f'{get_column_letter(start_col_ca+1)}{start_row_ca+1}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+1}'] = 'Level of Mapping'
    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+1}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+2, end_row=start_row_ca+3, end_column=start_col_ca+2)
    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+2}'] = 'Affinity'
    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+2}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca, start_column=start_col_ca+3, end_row=start_row_ca, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca}'] = 'Attainment % in'
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+3, end_row=start_row_ca+1, end_column=start_col_ca+8)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+1}'] = 'Direct'
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+1}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+3, end_row=start_row_ca+2, end_column=start_col_ca+4)
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+2}'] = 'University(SEE)'
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+2}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+3}'] = 'Attainment'
    aw[f'{get_column_letter(start_col_ca+3)}{start_row_ca+3}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+4)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    aw[f'{get_column_letter(start_col_ca+4)}{start_row_ca+3}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+5, end_row=start_row_ca+2, end_column=start_col_ca+6)
    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+2}'] = 'Internal(CIE)'
    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+2}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+3}'] = 'Attainment'
    aw[f'{get_column_letter(start_col_ca+5)}{start_row_ca+3}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+6)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    aw[f'{get_column_letter(start_col_ca+6)}{start_row_ca+3}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+7, end_row=start_row_ca+2, end_column=start_col_ca+8)
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+2}'] = '="Weighted Level of Attainment (" & B16 & " SEE + " & B15 & " CIE)"'
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+2}'].font = Font(bold=True)
    aw.row_dimensions[start_row_ca+2].height = 52

    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+3}'] = 'Attainment'
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+3}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+9, end_row=start_row_ca+1, end_column=start_col_ca+10)
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+1}'] = 'Indirect'
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+1}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+9, end_row=start_row_ca+3, end_column=start_col_ca+9)
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+2}'] = 'Attainment'
    aw[f'{get_column_letter(start_col_ca+9)}{start_row_ca+2}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+2, start_column=start_col_ca+10, end_row=start_row_ca+3, end_column=start_col_ca+10)
    aw[f'{get_column_letter(start_col_ca+10)}{start_row_ca+2}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    aw[f'{get_column_letter(start_col_ca+10)}{start_row_ca+2}'].font = Font(bold=True)

    aw.merge_cells(start_row=start_row_ca+1, start_column=start_col_ca+11, end_row=start_row_ca+2, end_column=start_col_ca+12)
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+1}'] = 'Final Weighted CO Attainment (80% Direct + 20% Indirect)'
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+1}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+3}'] = 'Attainment'
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+3}'].font = Font(bold=True)

    aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}'] = 'Level Of Attainment (0-40 --> 1, 40-60 ---> 2, 60-100---> 3)'
    aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}'].font = Font(bold=True)


    aw.column_dimensions[get_column_letter(start_col_ca)].width = 17.22
    aw.column_dimensions[get_column_letter(start_col_ca+1)].width = 9.33
    aw.column_dimensions[get_column_letter(start_col_ca+2)].width = 15.56
    aw.column_dimensions[get_column_letter(start_col_ca+3)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+4)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+5)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+6)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+7)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+8)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+9)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+10)].width = 12
    aw.column_dimensions[get_column_letter(start_col_ca+11)].width = 13
    aw.column_dimensions[get_column_letter(start_col_ca+12)].width = 12
    
    for row in aw.iter_rows(min_row=start_row_ca, max_row=aw.max_row, min_col=start_col_ca, max_col=start_col_ca+12):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
    aw[f'{get_column_letter(start_col_ca+2)}{start_row_ca+2}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+2}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+7)}{start_row_ca+3}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+8)}{start_row_ca+3}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+1}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+11)}{start_row_ca+3}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
    aw[f'{get_column_letter(start_col_ca+12)}{start_row_ca+3}'].fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')


    #================================================================================================================================================================

    start=start_row_ca+3
    interval=16
    rowindex=1
    for i in range(1, (data["Number_of_COs"]+1)):
        start+=1
        aw.merge_cells(start_row=start, start_column=start_col_ca, end_row=start+interval, end_column=start_col_ca)
        aw.cell(row=start, column=start_col_ca).value = "CO"+str(i)
        aw.cell(row=start, column=start_col_ca).font = Font(bold=True)
        aw.cell(row=start, column=start_col_ca).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if i%2==0:
            aw.cell(row=start, column=start_col_ca).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
        else:
            aw.cell(row=start, column=start_col_ca).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')


        index=1
        for j in range(start, start+interval+1):
            #print out COPO mapping
            aw.cell(row=j, column=start_col_ca+1).value =f'={get_column_letter(index+4)}2'
            aw.cell(row=j, column=start_col_ca+1).alignment = Alignment(horizontal='center', vertical='center')
            aw.cell(row=j, column=start_col_ca+1).border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
            aw.cell(row=j, column=start_col_ca+2).value = f'={get_column_letter(index+4)}{i+2}'
            aw.cell(row=j, column=start_col_ca+2).alignment = Alignment(horizontal='center', vertical='center')
            aw.cell(row=j, column=start_col_ca+2).border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            
            if index%2==0:
                aw.cell(row=j, column=start_col_ca+1).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
                aw.cell(row=j, column=start_col_ca+2).fill = PatternFill(start_color='c4d79b', end_color='c4d79b', fill_type='solid')
            else:
                aw.cell(row=j, column=start_col_ca+1).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
                aw.cell(row=j, column=start_col_ca+2).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            index+=1

        for k in range(start_col_ca+3, start_col_ca+13):
            aw.merge_cells(start_row=start, start_column=k, end_row=start+interval, end_column=k)
            aw.cell(row=start, column=k).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if k%2==0:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='b8cce4', end_color='b8cce4', fill_type='solid')
            else:
                aw.cell(row=start, column=k).fill = PatternFill(start_color='dce6f1', end_color='dce6f1', fill_type='solid')
    
        internal_components_num=0
        external_components_num=0
        for component_name in Component_Details.keys():
            if component_name[-1]=="I":
                internal_components_num+=1
            else:
                external_components_num+=1
        col=(data["Number_of_COs"]*internal_components_num) + (1*internal_components_num) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=start_col_ca+3).value=f'={data["Section"]}_Internal_Components!{get_column_letter(col)}{row}'

        aw.cell(row=start, column=start_col_ca+4).value=f'=IF(AND({get_column_letter(start_col_ca+3)}{start}>0,{get_column_letter(start_col_ca+3)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+3)}{start}>=40,{get_column_letter(start_col_ca+3)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+3)}{start}>=60,{get_column_letter(start_col_ca+3)}{start}<=100),3,"0")))'

        col=(data["Number_of_COs"]*external_components_num) + (1*external_components_num) + 2 + rowindex
        row=(6 + data["Number_of_Students"] + 5)
        aw.cell(row=start, column=start_col_ca+5).value=f'={data["Section"]}_External_Components!{get_column_letter(col)}{row}'

        aw.cell(row=start, column=start_col_ca+6).value=f'=IF(AND({get_column_letter(start_col_ca+5)}{start}>0,{get_column_letter(start_col_ca+5)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+5)}{start}>=40,{get_column_letter(start_col_ca+5)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+5)}{start}>=60,{get_column_letter(start_col_ca+5)}{start}<=100),3,"0")))'

        SEE_attainment = f'{get_column_letter(start_col_ca+3)}{start}'
        CIE_attainment = f'{get_column_letter(start_col_ca+5)}{start}'
        calculation = f'{SEE_attainment}*(B16/100)+{CIE_attainment}*(B15/100)'

        formula = f'={calculation}'
        aw.cell(row=start, column=start_col_ca+7).value = formula

        aw.cell(row=start, column=start_col_ca+8).value=f'=IF(AND({get_column_letter(start_col_ca+7)}{start}>0,{get_column_letter(start_col_ca+7)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+7)}{start}>=40,{get_column_letter(start_col_ca+7)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+7)}{start}>=60,{get_column_letter(start_col_ca+7)}{start}<=100),3,"0")))'

        aw.cell(row=start, column=start_col_ca+9).value=f'=E{2+data["Number_of_COs"]+4+rowindex}'

        aw.cell(row=start, column=start_col_ca+10).value=f'=IF(AND({get_column_letter(start_col_ca+9)}{start}>0,{get_column_letter(start_col_ca+9)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+9)}{start}>=40,{get_column_letter(start_col_ca+9)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+9)}{start}>=60,{get_column_letter(start_col_ca+9)}{start}<=100),3,"0")))'

        direct_attainment = f'{get_column_letter(start_col_ca+7)}{start}'
        indirect_attainment = f'{get_column_letter(start_col_ca+9)}{start}'
        calculation = f'={direct_attainment}*(B17/100)+{indirect_attainment}*(B18/100)'
        formula = f'={calculation}'
        aw.cell(row=start, column=start_col_ca+11).value = formula

        aw.cell(row=start, column=start_col_ca+12).value=f'=IF(AND({get_column_letter(start_col_ca+11)}{start}>0,{get_column_letter(start_col_ca+11)}{start}<40),1,IF(AND({get_column_letter(start_col_ca+11)}{start}>=40,{get_column_letter(start_col_ca+11)}{start}<60),2,IF(AND({get_column_letter(start_col_ca+11)}{start}>=60,{get_column_letter(start_col_ca+11)}{start}<=100),3,"0")))'

        rowindex+=1
        start=start+interval

    for row in aw.iter_rows(min_row=start_row_ca+4, max_row=aw.max_row, min_col=start_col_ca, max_col=start_col_ca+12):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            

    #================================================================================================================================================================
    current_row = aw.max_row+4
    current_col = start_col_ca
    aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=17+current_col)
    aw.cell(row=current_row, column=current_col).value = "Weighted PO/PSO Attainment Contribution"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    

    current_row+=1
    aw.cell(row=current_row, column=current_col).value = "COs\\POs"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True, color="FFFFFF")
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    for po in range(1,12+1):
        aw[f"{get_column_letter(po+current_col)}{current_row}"]=f"PO{po}"
        aw[f"{get_column_letter(po+current_col)}{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"{get_column_letter(po+current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(po+current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(po+current_col)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        
    
    for pso in range(1,6):
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"]=f"PSO{pso}"
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
        

    current_row+=1

    start=start_row_ca+3
    interval=17

    current_col=start_col_ca
    for co in range(1,data["Number_of_COs"]+1):
        aw[f"{get_column_letter(current_col)}{current_row}"]=f"CO{co}"
        aw[f"{get_column_letter(current_col)}{current_row}"].font = Font(bold=True, color="FFFFFF")
        aw[f"{get_column_letter(current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
        aw[f"{get_column_letter(current_col)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

        for po in range(1,12+1):
            aw[f"{get_column_letter(po+current_col)}{current_row}"]=f'={get_column_letter(start_col_ca+2)}{start+po}*{get_column_letter(start_col_ca+12)}{start+1}'
            aw[f"{get_column_letter(po+current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            aw[f"{get_column_letter(po+current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
            
            

        for pso in range(1,6):
            aw[f"{get_column_letter(12+current_col+pso)}{current_row}"]=f'={get_column_letter(start_col_ca+2)}{start+12+pso}*{get_column_letter(start_col_ca+12)}{start+1}'
            aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))

        current_row+=1
        start+=interval

    current_col=1
    aw.cell(row=current_row, column=current_col).value = "Academic Year"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).value = data["Academic_year"]
    aw.cell(row=current_row+1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row+1, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    aw.cell(row=current_row+1, column=current_col).font = Font(bold=True, color="FFFFFF")

    current_col+=1
    aw.cell(row=current_row, column=current_col).value = "Semester"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).value = data["Semester"]
    aw.cell(row=current_row+1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row+1, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    aw.cell(row=current_row+1, column=current_col).font = Font(bold=True, color="FFFFFF")
    current_col+=1

    aw.cell(row=current_row, column=current_col).value = "Subject Name"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).value = data["Subject_Name"]
    aw.cell(row=current_row+1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row+1, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    aw.cell(row=current_row+1, column=current_col).fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')
    aw.cell(row=current_row+1, column=current_col).font = Font(bold=True, color="FFFFFF")
    current_col+=1

    aw.cell(row=current_row, column=current_col).value = "Subject Code"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    
    current_col+=1





    aw.merge_cells(start_row=current_row, start_column=current_col, end_row=current_row, end_column=16+current_col)
    aw.cell(row=current_row, column=current_col).value = "Final Ratio"
    aw.cell(row=current_row, column=current_col).font = Font(bold=True)
    aw.cell(row=current_row, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
    aw.cell(row=current_row, column=current_col).fill = PatternFill(start_color='E6BA62', end_color='E6BA62', fill_type='solid')
    aw.cell(row=current_row, column=current_col).border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    

    current_row+=1
    current_col=4
    aw[f"{get_column_letter(current_col)}{current_row}"]=f"{data['Subject_Code']}"
    aw[f"{get_column_letter(current_col)}{current_row}"].font = Font(bold=True, color="FFFFFF")
    aw[f"{get_column_letter(current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
    aw[f"{get_column_letter(current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
    aw[f"{get_column_letter(current_col)}{current_row}"].fill = PatternFill(start_color='4bacc6', end_color='4bacc6', fill_type='solid')

    for po in range(1,12+1):
        main_formula = f'SUM({get_column_letter(po+current_col)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(po+current_col)}{current_row-2})/(SUM({get_column_letter(4+po)}3:{get_column_letter(4+po)}{data["Number_of_COs"]+2}))'
        complete_formula = f'=IF(AND(SUM({get_column_letter(po+current_col)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(po+current_col)}{current_row-2})>0, SUM({get_column_letter(4+po)}3:{get_column_letter(4+po)}{data["Number_of_COs"]+2})>0), {main_formula}, 0)'
        aw[f"{get_column_letter(po+current_col)}{current_row}"] = complete_formula
        aw[f"{get_column_letter(po+current_col)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(po+current_col)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    for pso in range(1,6):
        main_formula = f'SUM({get_column_letter(12+current_col+pso)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(12+current_col+pso)}{current_row-2})/(SUM({data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}3:{data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}{data["Number_of_COs"]+2}))'
        complete_formula = f'=IF(AND(SUM({get_column_letter(12+current_col+pso)}{current_row-1-data["Number_of_COs"]}:{get_column_letter(12+current_col+pso)}{current_row-2})>0, SUM({data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}3:{data["Section"]}_Input_Details!{get_column_letter(12+4+pso)}{data["Number_of_COs"]+2})>0), {main_formula}, 0)'
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"] = complete_formula
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
        aw[f"{get_column_letter(12+current_col+pso)}{current_row}"].border = Border(left=Side(border_style='thin', color='000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    current_row+=1

    return aw